"""
app_helpers.py — pure-Python business logic ported from the legacy Flask app.

Every route handler from the old `bom-tool/app.py` lives here as a plain
function. The Flask `request` object is gone — handlers now take explicit
parameters. The accompanying FastAPI routers in `app/routes/*` are thin
wrappers that parse the FastAPI Request into these parameters and call back
into this module.

Migrated in Phase 3.1 of the refactor (docs/MIGRATION_BOM_FLASK_TO_FASTAPI.md
slices 1-11). The original Flask `app.py` and `wsgi.py` are gone; this module
plus the FastAPI app are the canonical implementation.
"""
from __future__ import annotations

from werkzeug.utils import secure_filename
import pandas as pd
import os, threading, time, re, math, json, io, csv
from concurrent.futures import ThreadPoolExecutor, wait as _wait_futures
from datetime import datetime, timezone
from dotenv import load_dotenv
from urllib.parse import quote_plus
from typing import Any, Dict

from fastapi import HTTPException

# Load secrets before importing vendor modules; several adapters read env vars at import time.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"), override=False)

from database import (init_db, get_db, create_rfq, add_vendor, get_all_vendors,
                      add_contract_manufacturer, get_all_contract_manufacturers,
                      delete_contract_manufacturer, save_price_history,
                      get_price_history, get_price_trend, get_cheapest_ever,
                      upsert_rfq_vendor_db, get_rfq_vendor_db_masked,
                      get_rfq_vendor_actual, clear_rfq_vendor_db,
                      match_vendors_to_bom, get_vendors_by_rfq_type,
                      category_to_rfq_type)
from Mouser_fetch import get_mouser_price
from Digikey_fetch import get_digikey_token, get_digikey_price
from element14_fetch import get_element14_price
from lcsc_fetch import get_lcsc_price
from arrow_fetch import get_arrow_price
from indian_stores_fetch import get_indian_best_price
from Email_Sender import send_rfq_email, send_cdp_rfq, send_mechanical_rfq
from gmail_parser import check_rfq_replies
from hsn_lookup import get_hsn_bulk, infer_hsn_local, infer_manufacturer_local
from alt_component import get_alternatives_bulk
from custom_duty import calculate_import_duty, calculate_bom_duties, get_duty_rates
from cybex_lookup import verify_hsn
from eol_fetch import get_eol_bulk
from shopping_search import shopping_fallback_search, get_google_shopping_results
try:
    from vendor_pricing.adapters.electronicscomp_scraper import ElectronicsCompScraper
    from vendor_pricing.adapters.evelta_scraper import EveltaScraper
    from vendor_pricing.adapters.flyrobo_scraper import FlyroboScraper
    from vendor_pricing.adapters.robocraze_scraper import RobocrazeScraper
    from vendor_pricing.adapters.robokits_scraper import RobokitsScraper
    from vendor_pricing.adapters.robu_scraper import RobuScraper
    from vendor_pricing.adapters.sharvi_scraper import SharviScraper
    from vendor_pricing.adapters.tenettech_scraper import TenettechScraper
    from vendor_pricing.services.vendor_aggregator import VendorAggregatorService
except Exception:
    VendorAggregatorService = None
    RobuScraper = EveltaScraper = ElectronicsCompScraper = RobokitsScraper = FlyroboScraper = RobocrazeScraper = None
    SharviScraper = TenettechScraper = None

import sys, logging
logging.basicConfig(
    stream=sys.stdout,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    force=True,
)
log = logging.getLogger("bom_tool")

# ── Filesystem prep — keeps RFQ uploads, replies and project files on disk. ──
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs("replies", exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, "project_files"), exist_ok=True)

# Ensure DB tables exist on every startup (works with Gunicorn too)
try:
    init_db()
except Exception as _db_err:
    print(f"[startup] DB init warning: {_db_err}", flush=True)

# ── DigiKey token: module-level cache so token fetch (5s timeout) happens
# only ONCE per server lifetime, not once per bulk-pricing request. ──────────
_cached_dk_token: str | None = None
_cached_dk_token_lock = threading.Lock()

def get_dk_token() -> str | None:
    """Return a cached DigiKey token; refresh only when expired or missing."""
    global _cached_dk_token
    # Digikey_fetch already has its own TTL cache internally — just call it.
    # We wrap it here with a fast short-circuit: if we already have a token,
    # return immediately without spawning a thread.
    try:
        # get_digikey_token() is fast when cache is warm (immediate return).
        # Only slow (~5s) on first call or after expiry — handled by caller's timeout.
        return get_digikey_token()
    except Exception:
        return None

FAST_ENDPOINT_TIMEOUT = float(os.getenv("FAST_ENDPOINT_TIMEOUT_SECONDS", "75"))


class _FlaskShim:
    """Minimal Flask-request shim that handlers receive instead of Flask's
    thread-local `request`.

    The FastAPI route layer (app/routes/*) builds one of these per inbound
    request from the typed FastAPI dependencies (UploadFile, Form, Query,
    Body) and passes it to the migrated handler. This kept the diff small:
    handler bodies still read `request.json`, `request.form`, `request.args`,
    `request.files`, `request.method` exactly as they did under Flask, but
    the global is gone.

    Only the subset of the Flask API that the migrated handlers actually
    touch is implemented. If a future handler needs something more, add it
    here, not by re-introducing Flask.
    """

    class _FileShim:
        """Mimics a werkzeug FileStorage so handlers can call .read()/.save()/.filename."""
        def __init__(self, upload):
            self._upload = upload
            self.filename = getattr(upload, "filename", "") or ""
            self.mimetype = getattr(upload, "content_type", "") or ""
            self._cached_bytes = None

        def read(self):
            if self._cached_bytes is None:
                try:
                    self._upload.file.seek(0)
                except Exception:
                    pass
                self._cached_bytes = self._upload.file.read()
            return self._cached_bytes

        def save(self, dst):
            data = self.read()
            with open(dst, "wb") as fh:
                fh.write(data)

    class _Files:
        def __init__(self, files):
            self._files = {}
            for k, v in (files or {}).items():
                self._files[k] = _FlaskShim._FileShim(v) if v is not None else None

        def __contains__(self, key):
            return key in self._files and self._files[key] is not None

        def __getitem__(self, key):
            return self._files[key]

        def get(self, key, default=None):
            v = self._files.get(key)
            return v if v is not None else default

        def getlist(self, key):
            f = self._files.get(key)
            return [f] if f else []

    class _DictWrap:
        def __init__(self, d):
            self._d = d or {}

        def get(self, key, default=None):
            v = self._d.get(key, default)
            return v if v is not None else default

        def __contains__(self, key):
            return key in self._d

        def __getitem__(self, key):
            return self._d[key]

    def __init__(
        self,
        *,
        method: str = "GET",
        json_body: Any = None,
        args: dict | None = None,
        form: dict | None = None,
        files: dict | None = None,
        host_url: str = "",
        headers: dict | None = None,
    ):
        self.method = method.upper()
        self.json = json_body
        self.args = self._DictWrap(args or {})
        self.form = self._DictWrap(form or {})
        self.files = self._Files(files or {})
        self.host_url = host_url or ""
        self.headers = headers or {}

    def get_json(self, silent: bool = False):
        return self.json


# ── Backwards-compat shims ─────────────────────────────────────────────────
# Handler bodies still call jsonify(x), send_file(...), render_template(...).
# Under FastAPI these become plain Python returns; the router shim wraps them.

def jsonify(data: Any) -> Any:
    """Compat shim — handlers still call jsonify(); FastAPI returns the dict."""
    return data


def send_file(buf, *, as_attachment: bool = False, download_name: str = "", mimetype: str = ""):
    """Compat shim — small descriptor that the FastAPI route layer converts
    into a StreamingResponse with the right headers.
    """
    return {
        "__send_file__": True,
        "buffer": buf,
        "as_attachment": as_attachment,
        "download_name": download_name,
        "mimetype": mimetype,
    }


def render_template(name: str, **ctx):
    """Compat shim — handlers only call this from the `/` index route, which
    we keep at FastAPI top-level in main.py. Defined here so the import line
    in this module doesn't break."""
    return {"__template__": name, "context": ctx}


class TemplateNotFound(Exception):
    """Compat — index route catches this. Kept identical name."""

FEATURE_STATUS = {
    "ai_bom_classification": "active",
    "ambiguous_review": "active",
    "manual_bom_entry": "active",
    "project_file_upload": "active",
    "file_versioning": "active",
    "assembly_fatp_rfq": "active",
    "alternative_live_pricing": "active_for_visible_alternative_mpn",
    "hsn_manual_approval": "active",
    "landed_cost_cif": "active",
    "quote_l1_l2_l3": "active",
    "excel_export": "active",
    "pdf_export": "printable_html_active",
    "negotiation_loop": "workflow_active_email_manual",
    "factory_match_score": "rule_based_active",
    "scheduling_calendar": "request_capture_active_calendar_api_pending",
    "production_tracking": "active",
    "whatsapp_bom": "pending_whatsapp_api",
    "official_hsn_api": "pending_provider_api",
    "lcsc_arrow_indian_stores": "active_lcsc_arrow_robu_evelta",
}

# ── Cache ──────────────────────────────────────────────────────
_price_cache = {}
_cache_ttl   = 3600
_cache_lock  = threading.Lock()
_vendor_pricing_service = None
_vendor_pricing_lock = threading.Lock()
_PRICE_CACHE_VERSION = "pricing-v6-raw-supplier-prices"
OVERSEAS_LANDED_MULTIPLIER = float(os.getenv("OVERSEAS_LANDED_MULTIPLIER", "1.5"))
OVERSEAS_LANDED_SUPPLIERS = {"mouser", "digikey", "element14", "arrow"}

def _cache_key(mpn, description=""):
    mpn_key = str(mpn or "").strip().upper()
    desc_key = re.sub(r"\s+", " ", str(description or "").strip().lower())[:120]
    return f"{_PRICE_CACHE_VERSION}|{mpn_key}|{desc_key}"

def _get_cached(mpn, description=""):
    with _cache_lock:
        e = _price_cache.get(_cache_key(mpn, description))
        if e:
            data, ts = e
            if time.time() - ts < _cache_ttl: return data
            del _price_cache[_cache_key(mpn, description)]
    return None

def _set_cached(mpn, data, description=""):
    with _cache_lock:
        _price_cache[_cache_key(mpn, description)] = (data, time.time())

# ── Helpers ────────────────────────────────────────────────────
def to_float(val):
    try: return float(str(val).replace('INR','').replace('₹','').replace(',','').strip())
    except: return None

def safe_int(val, default=1):
    try:
        v = float(str(val).strip())
        if math.isnan(v) or math.isinf(v): return default
        return int(v)
    except: return default

def positive_int(val, default=1):
    value = safe_int(val, default)
    return value if value > 0 else default

def clean_nan(val, default=""):
    if val is None: return default
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)): return default
    return val

def clean_comp(comp):
    return {k: (clean_nan(v,"") if not isinstance(v,int) else v) for k,v in comp.items()}

DNP_QTY_MARKERS = {"DNP", "DNF", "NM", "NP", "NO MOUNT", "NO FIT", "DO NOT FIT"}
DNP_TEXT_PATTERNS = (
    r"\bDNP\b", r"\bDNF\b", r"\bDO\s+NOT\s+POPULATE\b", r"\bDO\s+NOT\s+FIT\b",
    r"\bDO\s+NOT\s+PLACE\b", r"\bNO\s+MOUNT\b", r"\bNO\s+STUFF\b",
    r"\bNOT\s+POPULATED\b", r"\bNOT\s+FITTED\b"
)

def is_dnp_component(qty_value="", status_value="", description="", item_name=""):
    """Conservatively detect do-not-populate rows without eating valid BOP lines."""
    qty_text = str(clean_nan(qty_value, "")).strip().upper()
    status_text = str(clean_nan(status_value, "")).strip().upper()
    if qty_text in DNP_QTY_MARKERS or status_text in DNP_QTY_MARKERS:
        return True
    text = f"{clean_nan(description, '')} {clean_nan(item_name, '')}".upper()
    return any(re.search(pattern, text) for pattern in DNP_TEXT_PATTERNS)

def row_quantity(row, default=1):
    return positive_int(
        row.get("quantity", row.get("Quantity", row.get("Qty", default))),
        default,
    )

def file_ext(filename):
    return os.path.splitext(str(filename or "").lower())[1]

def is_csv_file(filename, mimetype=""):
    ext = file_ext(filename)
    return ext == ".csv" or "csv" in str(mimetype or "").lower()

def is_spreadsheet_file(filename, mimetype=""):
    ext = file_ext(filename)
    return ext in (".xlsx", ".xls", ".csv") or is_csv_file(filename, mimetype)

def price_for_quantity(info, qty):
    """Pick the best unit price for the requested component quantity.

    MOQ remains informational. Price selection follows distributor price breaks where available:
    highest break quantity <= requested quantity wins.
    """
    qty = safe_int(qty, 1)
    base_price = to_float((info or {}).get("price"))
    breaks = (info or {}).get("price_breaks") or []
    parsed = []
    for tier in breaks:
        try:
            tier_qty = int(tier.get("qty") or tier.get("Quantity") or tier.get("BreakQuantity") or 1)
        except (TypeError, ValueError):
            tier_qty = 1
        tier_price = to_float(tier.get("price") or tier.get("Price") or tier.get("UnitPrice"))
        if tier_price:
            parsed.append((tier_qty, tier_price))
    if not parsed:
        return base_price
    parsed.sort(key=lambda x: x[0])
    selected = parsed[0][1]
    for tier_qty, tier_price in parsed:
        if tier_qty <= qty:
            selected = tier_price
        else:
            break
    return selected

def price_tier_for_quantity(info, qty):
    qty = safe_int(qty, 1)
    breaks = (info or {}).get("price_breaks") or []
    parsed = []
    for tier in breaks:
        tier_qty = safe_int(tier.get("qty") or tier.get("Quantity") or tier.get("BreakQuantity") or 1, 1)
        tier_price = to_float(tier.get("price") or tier.get("Price") or tier.get("UnitPrice"))
        if tier_price:
            parsed.append((tier_qty, tier_price, tier))
    if not parsed:
        return None
    parsed.sort(key=lambda x: x[0])
    selected = parsed[0]
    for tier in parsed:
        if tier[0] <= qty:
            selected = tier
        else:
            break
    return selected[2]

def supplier_price_summary(info, qty):
    qty = safe_int(qty, 1)
    price_one = price_for_quantity(info, 1)
    price_qty = price_for_quantity(info, qty)
    tier = price_tier_for_quantity(info, qty) or {}
    return {
        "price_at_1": price_one,
        "price_at_requested_qty": price_qty,
        "requested_qty": qty,
        "requested_total": round(price_qty * qty, 2) if price_qty else None,
        "min_price_qty": safe_int((info or {}).get("min_price_qty") or (info or {}).get("moq") or 1, 1),
        "loose_available": bool((info or {}).get("loose_available", True)),
        "price_basis_package": tier.get("PackageType") or tier.get("package_type") or (info or {}).get("package_type") or "",
    }

def reprice_results_for_quantity(results, qty):
    repriced = {}
    for supplier, info in (results or {}).items():
        if not isinstance(info, dict):
            continue
        new_price = price_for_quantity(info, qty)
        repriced[supplier] = {**info, "price": new_price, "price_basis_qty": safe_int(qty, 1), **supplier_price_summary(info, qty)}
    return repriced

def read_csv_safely(path, **kwargs):
    kwargs.setdefault("sep", None)
    kwargs.setdefault("engine", "python")
    last_error = None
    for enc in ("utf-8-sig", "utf-16", "latin1"):
        try:
            return pd.read_csv(path, encoding=enc, **kwargs)
        except (UnicodeError, pd.errors.ParserError) as exc:
            last_error = exc
    if last_error:
        raise last_error
    return pd.read_csv(path, **kwargs)

def get_vendor_pricing_service():
    """Lazy-load scraper aggregation for Indian online stores.

    LCSC and Arrow remain handled by the existing API-style functions above. This supplemental
    service focuses on stores without confirmed public developer APIs.
    """
    global _vendor_pricing_service
    if VendorAggregatorService is None:
        return None
    with _vendor_pricing_lock:
        if _vendor_pricing_service is None:
            adapters = [
                RobuScraper(),
                EveltaScraper(),
                ElectronicsCompScraper(),
                RobokitsScraper(),
                FlyroboScraper(),
                RobocrazeScraper(),
            ]
            # ── FIXED: include Sharvi (SMD passives) and Tenettech (Murata/TDK) ──
            if SharviScraper is not None:
                adapters.append(SharviScraper())
            if TenettechScraper is not None:
                adapters.append(TenettechScraper())
            _vendor_pricing_service = VendorAggregatorService(adapters=adapters)
        return _vendor_pricing_service

def fetch_supplemental_vendor_prices(mpn, manufacturer="", description="", qty=1):
    service = get_vendor_pricing_service()
    if not service:
        return []
    try:
        return service.fetch_all_vendors_for_part(
            manufacturer_part_number=mpn,
            manufacturer_name=manufacturer or None,
            description=description or None,
            quantity=safe_int(qty, 1),
        )
    except Exception as exc:
        print(f"Supplemental vendor pricing failed for {mpn}: {exc}")
        return []

def fetch_indian_vendors(request):
    body = request.json or {}
    comps = body.get("components", []) or []
    def fetch_one_indian(comp):
        mpn = str(comp.get("mpn") or comp.get("MPN") or "").strip()
        desc = str(clean_nan(comp.get("description", comp.get("Description", "")), ""))
        manufacturer = str(clean_nan(comp.get("manufacturer", comp.get("Manufacturer", "")), ""))
        qty = positive_int(comp.get("quantity", comp.get("Quantity", 1)), 1)
        if not mpn and not desc:
            return None

        # Pricing pipeline:
        #   Round 1  – MPN exact search on all 10 Indian store adapters (VendorAggregatorService)
        #   Round 2  – Description keyword search on all 10 adapters
        #   Fallback – shopping_fallback_search spec-based (for passives: resistors, caps, etc.)
        #
        # Run MPN scrape first (parallel, 6s max across all adapters), then add
        # shopping_fallback spec results as supplementary coverage.
        results = {}

        # ── Round 1 + 2: VendorAggregatorService (MPN then description) ──────
        all_candidates = fetch_supplemental_vendor_prices(
            mpn or desc,
            manufacturer=manufacturer,
            description=desc,
            qty=qty,
        )
        for candidate in all_candidates:
            confidence = safe_int(candidate.get("match_confidence", 0), 0)
            if confidence < 44:
                continue
            legacy = normalized_vendor_to_legacy_result(candidate, qty)
            if not legacy:
                continue
            vendor_key = str(candidate.get("vendor_name") or "Vendor").strip()
            if vendor_key not in results or confidence > safe_int(results[vendor_key].get("match_confidence", 0), 0):
                legacy["match_type"] = "MPN" if confidence >= 60 else "Keyword"
                if confidence < 60:
                    legacy["description_match_warning"] = "Keyword match – verify specs before buying"
                results[vendor_key] = legacy

        # ── Fallback: spec-based shopping search (passives, connectors, etc.) ─
        # Only run if Round 1+2 found nothing — avoids wasting time on parts
        # already priced by the full scraper.
        if not results and (desc or mpn):
            try:
                # ── FIXED: hard 12-second timeout so a slow/hanging store
                # never blocks the entire indian-vendors response ──
                _sfi_ex = ThreadPoolExecutor(max_workers=1)
                _sfi_f  = _sfi_ex.submit(
                    shopping_fallback_search, mpn, desc, manufacturer,
                    max_results=6,
                    include_jina=bool(os.getenv("JINA_API_KEY")),
                )
                _sfi_done, _ = _wait_futures([_sfi_f], timeout=18)
                _sfi_hits = _sfi_f.result() if _sfi_done else []
                _sfi_ex.shutdown(wait=False, cancel_futures=True)
                for hit in _sfi_hits:
                    legacy = shopping_hit_to_legacy_result(hit, qty)
                    if not legacy:
                        continue
                    vendor_key = str(hit.get("vendor_name") or "Shopping").strip()
                    legacy["match_type"] = "Spec"
                    legacy["description_match_warning"] = "Spec/Google result - verify specs before PO"
                    if (
                        vendor_key not in results
                        or safe_int(legacy.get("match_confidence"), 0) > safe_int(results[vendor_key].get("match_confidence"), 0)
                    ):
                        results[vendor_key] = legacy
            except Exception as exc:
                print(f"Shopping spec search failed for {mpn or desc}: {exc}")

        status = "found" if results else "not_found"
        best_supplier, best_price = best_supplier_from_results(results)
        return {
            "mpn": mpn,
            "description": desc,
            "manufacturer": manufacturer,
            "quantity": qty,
            "results": results,
            "best_supplier": best_supplier,
            "best_price": best_price,
            "total_cost": round(best_price * qty, 2) if best_price else None,
            "indian_vendor_status": status,
        }
    max_workers = min(max(len(comps), 1), 12)
    rows = []
    ex = ThreadPoolExecutor(max_workers=max_workers)
    try:
        futures = {ex.submit(fetch_one_indian, comp): comp for comp in comps}
        done, pending = _wait_futures(futures.keys(), timeout=min(FAST_ENDPOINT_TIMEOUT, 32))
        for fut in pending:
            fut.cancel()
            comp = futures[fut]
            rows.append({
                "mpn": str(comp.get("mpn") or comp.get("MPN") or "").strip(),
                "description": str(clean_nan(comp.get("description", comp.get("Description", "")), "")),
                "manufacturer": str(clean_nan(comp.get("manufacturer", comp.get("Manufacturer", "")), "")),
                "quantity": positive_int(comp.get("quantity", comp.get("Quantity", 1)), 1),
                "results": {},
                "best_supplier": None,
                "best_price": None,
                "total_cost": None,
                "indian_vendor_status": "timeout_rfq",
            })
        for fut in done:
            try:
                row = fut.result()
                if row:
                    rows.append(row)
            except Exception as exc:
                print(f"Indian vendor row failed: {exc}")
    finally:
        ex.shutdown(wait=False, cancel_futures=True)
    return safe_jsonify(rows)

def demo_coverage_search(request):
    """One-line deep search for demo coverage.

    Intentionally avoids the full adapter pool — searches one BOM line at a
    time with hard time budgets to keep latency predictable for demos.
    """
    body = request.json or {}
    comp = clean_comp((body.get("component") or body.get("components") or [{}])[0] if isinstance(body.get("components"), list) else body.get("component", {}))
    mpn = str(comp.get("MPN") or comp.get("mpn") or "").strip()
    desc = str(clean_nan(comp.get("Description", comp.get("description", "")), ""))
    manufacturer = str(clean_nan(comp.get("Manufacturer", comp.get("manufacturer", "")), ""))
    qty = positive_int(comp.get("Quantity", comp.get("quantity", 1)), 1)
    original_qty = positive_int(comp.get("original_quantity", qty), qty)
    pcb_qty = positive_int(comp.get("pcb_quantity", 1), 1)
    if not mpn and not desc:
        return jsonify({"error": "MPN or description required"}), 400

    results = {}
    raw_sources = {}

    def run_lcsc():
        try:
            return get_lcsc_price(mpn)
        except Exception:
            return None

    def run_shopping():
        try:
            return shopping_fallback_search(
                mpn,
                desc,
                manufacturer,
                max_results=6,
                include_jina=bool(os.getenv("JINA_API_KEY")),
            )
        except Exception as exc:
            print(f"Demo coverage shopping failed for {mpn or desc}: {exc}")
            return []

    ex = ThreadPoolExecutor(max_workers=2)
    try:
        futures = {ex.submit(run_lcsc): "lcsc", ex.submit(run_shopping): "shopping"}
        done, pending = _wait_futures(futures.keys(), timeout=18)
        for fut in pending:
            fut.cancel()
        for fut in done:
            key = futures[fut]
            try:
                raw_sources[key] = fut.result()
            except Exception:
                raw_sources[key] = None
    finally:
        ex.shutdown(wait=False, cancel_futures=True)

    lcsc = raw_sources.get("lcsc")
    if lcsc and to_float(lcsc.get("price")):
        lcsc = apply_overseas_landed_price("lcsc", lcsc)
        price = price_for_quantity(lcsc, qty) or to_float(lcsc.get("price"))
        results["lcsc"] = {
            "price": price,
            "base_price": to_float(lcsc.get("price")),
            "price_basis_qty": qty,
            "price_breaks": lcsc.get("price_breaks") or [],
            **supplier_price_summary(lcsc, qty),
            "stock": str(clean_nan(lcsc.get("stock"), "N/A")),
            "lead_time": str(clean_nan(lcsc.get("lead_time"), "N/A")),
            "moq": clean_nan(lcsc.get("moq"), 1),
            "min_price_qty": clean_nan(lcsc.get("min_price_qty"), lcsc.get("moq", 1)),
            "url": str(clean_nan(lcsc.get("url"), "")),
            "product_title": str(clean_nan(lcsc.get("description"), "")),
            "match_confidence": safe_int(lcsc.get("match_confidence"), 72),
            "source_type": "lcsc_demo_coverage",
        }

    _VENDOR_KEY_NORM = {"LCSC": "lcsc", "Lcsc": "lcsc"}
    for hit in raw_sources.get("shopping") or []:
        if safe_int(hit.get("match_confidence", 0), 0) < 40:
            continue
        legacy = shopping_hit_to_legacy_result(hit, qty)
        if not legacy:
            continue
        vendor_key = _VENDOR_KEY_NORM.get(str(hit.get("vendor_name") or "Online").strip(),
                                          str(hit.get("vendor_name") or "Online").strip())
        legacy["match_type"] = "Spec"
        legacy["description_match_warning"] = "Demo coverage result - verify specs before PO"
        existing = results.get(vendor_key)
        if not existing or safe_int(legacy.get("match_confidence"), 0) > safe_int(existing.get("match_confidence"), 0):
            results[vendor_key] = legacy

    mfr = manufacturer or infer_manufacturer_local(mpn, desc)
    hsn_local = infer_hsn_local(mpn, desc, mfr)
    hsn = str(hsn_local.get("hsn_code", "")).strip()
    b, bp = best_supplier_from_results(results)
    row = {
        "mpn": mpn,
        "description": desc,
        "manufacturer": mfr,
        "hsn_code": hsn if is_valid_hsn(hsn) else "N/A",
        "hsn_desc": hsn_local.get("hsn_desc", ""),
        "hsn_confidence": hsn_local.get("confidence", "Medium"),
        "hsn_source": "Local Rule Proposed",
        "hsn_warning": "" if is_valid_hsn(hsn) else "HSN not verified yet",
        "hsn_reference_url": cybex_hsn_url(hsn, desc),
        "datasheet_url": extract_datasheet_from_sources(lcsc or {}),
        "results": results,
        "alternatives": [],
        "quantity": qty,
        "original_quantity": original_qty,
        "pcb_quantity": pcb_qty,
        "best_supplier": b,
        "best_price": bp,
        "total_cost": round(bp * qty, 2) if bp else None,
        "cached": False,
        "pricing_status": "coverage_found" if bp else "coverage_not_found",
    }
    return safe_jsonify(row)

def normalized_vendor_to_legacy_result(candidate, qty=1):
    price = to_float(candidate.get("unit_price"))
    if not price:
        return None
    stock = candidate.get("stock")
    # Only skip if the store explicitly says out-of-stock.
    # Scrapers often can't detect stock status (in_stock defaults to False when
    # stock text is empty) — don't discard those results.
    stock_lower = str(stock or "").lower()
    if any(phrase in stock_lower for phrase in ("out of stock", "sold out", "unavailable")):
        return None
    return {
        "price": price,
        "base_price": price,
        "price_basis_qty": safe_int(qty, 1),
        "price_breaks": candidate.get("price_breaks") or [],
        **supplier_price_summary({"price": price, "price_breaks": candidate.get("price_breaks") or []}, qty),
        "stock": str(clean_nan(stock, "N/A")),
        "lead_time": str(clean_nan(candidate.get("lead_time"), "N/A")),
        "moq": clean_nan(candidate.get("moq"), 1),
        "url": str(clean_nan(candidate.get("product_url"), "")),
        "product_title": str(clean_nan(candidate.get("product_title"), "")),
        "manufacturer": str(clean_nan(candidate.get("manufacturer_name"), "")),
        "source_type": candidate.get("source_type"),
        "match_confidence": candidate.get("match_confidence", 0),
        "vendor_part_number": candidate.get("vendor_part_number"),
    }

def apply_overseas_landed_price(supplier, info):
    """Annotate overseas distributor rows without changing displayed supplier price.

    Supplier columns must show the exact unit price returned by the vendor/API.
    A landed multiplier can be used later for duty/CIF estimates, but it must not
    overwrite `price`, `base_price`, or `price_breaks`; otherwise users see
    inflated vendor prices such as 10.88 becoming 16.32.
    """
    supplier_key = str(supplier or "").strip().lower()
    if supplier_key not in OVERSEAS_LANDED_SUPPLIERS or not isinstance(info, dict):
        return info
    multiplier = OVERSEAS_LANDED_MULTIPLIER
    if multiplier <= 0 or abs(multiplier - 1) < 0.0001:
        return info
    adjusted = dict(info)
    for key in ("price", "base_price", "price_at_1", "price_at_requested_qty"):
        value = to_float(adjusted.get(key))
        if value:
            adjusted[f"estimated_landed_{key}"] = round(value * multiplier, 4)
    requested_landed = to_float(adjusted.get("estimated_landed_price_at_requested_qty"))
    if requested_landed:
        adjusted["estimated_landed_requested_total"] = round(
            requested_landed * safe_int(adjusted.get("requested_qty"), 1), 2
        )
    adjusted["landed_multiplier"] = multiplier
    adjusted["landed_applied"] = False
    return adjusted

def shopping_hit_to_legacy_result(hit, qty=1):
    """Convert spec/google-style shopping hits into the dashboard supplier format."""
    price = to_float(hit.get("unit_price") or hit.get("price"))
    if not price:
        return None
    price_breaks = hit.get("price_breaks") or [{"qty": 1, "price": price}]
    return {
        "price": price_for_quantity({"price": price, "price_breaks": price_breaks}, qty) or price,
        "base_price": price,
        "price_basis_qty": safe_int(qty, 1),
        "price_breaks": price_breaks,
        **supplier_price_summary({"price": price, "price_breaks": price_breaks}, qty),
        "stock": str(clean_nan(hit.get("stock"), "Online / verify")),
        "lead_time": str(clean_nan(hit.get("lead_time"), "Online stock - verify checkout")),
        "moq": clean_nan(hit.get("moq"), 1),
        "min_price_qty": clean_nan(hit.get("moq"), 1),
        "url": str(clean_nan(hit.get("product_url") or hit.get("url"), "")),
        "product_title": str(clean_nan(hit.get("product_title"), "")),
        "manufacturer": str(clean_nan(hit.get("manufacturer_name"), "")),
        "source_type": hit.get("source_type") or "shopping_spec_match",
        "match_confidence": safe_int(hit.get("match_confidence"), 68),
        "search_query": str(clean_nan(hit.get("search_query"), "")),
        "loose_available": True,
    }

def best_supplier_from_results(results):
    best_name, best_price_value = None, None
    for supplier, info in (results or {}).items():
        if not isinstance(info, dict):
            continue
        price = to_float(info.get("price"))
        if price and (best_price_value is None or price < best_price_value):
            best_name, best_price_value = supplier, price
    return best_name, best_price_value

def extract_field(r, *fields):
    for f in fields:
        val = r.get(f,'')
        if val and str(val).strip().lower() not in ('none','n/a',''): return str(val).strip()
    return ''

def extract_mfr_desc(desc, mpn=""):
    if not desc: return ""
    desc = str(desc).strip()
    if " - " in desc:
        p = desc.split(" - ")[0].strip()
        if p.upper() != mpn.upper() and len(p) > 1: return p
    return ""

def extract_mfr_url(url):
    if not url: return ""
    m = re.search(r'/products/detail/([^/]+)/', str(url))
    return m.group(1).replace('-',' ').title() if m else ""

def extract_datasheet_from_sources(*sources):
    for source in sources:
        if not source:
            continue
        url = extract_field(source, "datasheet_url", "DataSheetUrl", "DatasheetUrl", "PrimaryDatasheet")
        if url:
            return url
    return ""

def apply_lifecycle_fallback(row):
    status = str(row.get("eol_status") or row.get("status") or "").strip().lower()
    if status and status not in ("unknown", "n/a", "-"):
        return row
    has_price = bool(row.get("best_price"))
    has_stock = False
    for info in (row.get("results") or {}).values():
        stock = str(info.get("stock", "")).lower()
        if stock and stock not in ("n/a", "none", "-", "0"):
            digits = "".join(ch for ch in stock if ch.isdigit())
            if not digits or safe_int(digits, 0) > 0 or "stock" in stock or "available" in stock:
                has_stock = True
                break
    if has_price or has_stock:
        row["eol_status"] = "Active"
        row["eol_risk"] = "low"
        row["eol_color"] = "green"
        row["eol_source"] = "Supplier availability"
        row["eol_note"] = "Live supplier price/stock found; verify PCN lifecycle for production release."
    return row

def safe_jsonify(data):
    """JSON-safe payload: strips NaN/Inf so FastAPI's encoder can return cleanly.

    Originally Flask returned a flask.Response with a hand-built body. Under
    FastAPI we just hand back the already-cleaned Python object; FastAPI's
    response encoder serialises it. Float NaN/Inf get coerced to None first
    so the encoder doesn't choke.
    """
    return json.loads(json.dumps(data, default=lambda x: None if (
        isinstance(x, float) and (math.isnan(x) or math.isinf(x))) else x))

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def ensure_prd_tables():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS project_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT,
            original_name TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            stream TEXT DEFAULT 'Project Docs',
            version INTEGER DEFAULT 1,
            size_bytes INTEGER DEFAULT 0,
            uploaded_at TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS hsn_approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mpn TEXT NOT NULL,
            hsn_code TEXT NOT NULL,
            hsn_desc TEXT DEFAULT '',
            source TEXT DEFAULT 'Manual Verified',
            approved_by TEXT DEFAULT 'user',
            approved_at TEXT NOT NULL,
            UNIQUE(mpn, hsn_code)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS manual_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT,
            stream TEXT DEFAULT 'BOP',
            vendor_name TEXT NOT NULL,
            mpn TEXT NOT NULL,
            unit_price REAL NOT NULL,
            moq INTEGER DEFAULT 1,
            lead_time TEXT DEFAULT '',
            certification TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS production_milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT,
            stage TEXT NOT NULL,
            owner TEXT DEFAULT '',
            status TEXT DEFAULT 'Pending',
            due_date TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            updated_at TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS negotiation_rounds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT,
            target_price REAL DEFAULT 0,
            round_no INTEGER DEFAULT 1,
            status TEXT DEFAULT 'Open',
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS scheduling_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT,
            cm_name TEXT DEFAULT '',
            preferred_date TEXT DEFAULT '',
            attendees TEXT DEFAULT '',
            status TEXT DEFAULT 'Calendar API Pending',
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS vendor_search_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_code TEXT DEFAULT '',
            part_name TEXT NOT NULL,
            item_type TEXT NOT NULL DEFAULT 'CDP',
            scale TEXT NOT NULL DEFAULT 'sample',
            vendor_name TEXT NOT NULL,
            website TEXT DEFAULT '',
            email TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            location TEXT DEFAULT 'India',
            supplies TEXT DEFAULT '',
            relevance INTEGER DEFAULT 0,
            why_trusted TEXT DEFAULT '',
            min_order TEXT DEFAULT '',
            website_reachable INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            UNIQUE(part_name, item_type, scale, vendor_name)
        )
    """)
    conn.commit(); conn.close()

def classify_project_file(name):
    n = str(name or '').lower()
    ext = os.path.splitext(n)[1]
    if ext in ('.step','.stp','.stl','.dxf','.dwg'): return 'Mechanical'
    if ext in ('.zip',) and ('gerber' in n or 'pcb' in n): return 'Electronics'
    if ext in ('.sch','.brd','.kicad_pcb','.dsn'): return 'Electronics'
    if ext in ('.xlsx','.xls','.csv'): return 'Production'
    if ext in ('.pdf','.docx','.doc'): return 'Project Docs'
    if ext in ('.jpg','.jpeg','.png'): return 'Design'
    return 'Project Docs'

def classify_bom_line(item):
    text = " ".join(str(item.get(k,'') or '') for k in ('MPN','mpn','part_name','Description','description','note')).lower()
    mpn = str(item.get('MPN') or item.get('mpn') or item.get('part_name') or '').strip()

    # Mechanical / custom / fabricated items must not leak into BOP.
    custom_words = [
        'custom','machined','mould','mold','enclosure','bracket','sheet metal','sheet-metal',
        'cnc','fabricated','extrusion','harness','rubber','mjf','sls','sla','fdm',
        '3d print','3d-printed','printed','standoff','spacer','cover','housing','chassis',
        'knob','shaft','gear','bushing','fixture','jig','gasket','panel','clip','latch',
        'hinge','spring','fastener','washer','nut','bolt','screw','metal','plastic',
        'insert','spacer','base','body','frame','carrier','guide','holder','mechanical','mech',
        'light pipe','top body','bottom body','faceplate','front cover','rear cover',
        'shell','bezel','trim'
    ]
    assembly_words = ['assembly','pcba','smt','fatp','testing','packaging','box build','through-hole','through hole']
    bop_words = ['resistor','capacitor','ic','connector','diode','transistor','module','ferrite','bead','switch','led','crystal','header','usb']

    if any(w in text for w in assembly_words):
        return force_cdp_if_mechanical(item, {'item_type':'ASSEMBLY','confidence':88,'reason':'Assembly/FATP keyword detected','ambiguous':False})

    if any(w in text for w in custom_words):
        return force_cdp_if_mechanical(item, {'item_type':'CDP','confidence':95,'reason':'Mechanical/custom manufacturing keyword detected','ambiguous':False})

    if is_valid_mpn(mpn) or any(w in text for w in bop_words):
        conf = 82 if is_valid_mpn(mpn) else 68
        return force_cdp_if_mechanical(item, {'item_type':'BOP','confidence':conf,'reason':'Standard purchasable component pattern','ambiguous':conf < 75})

    if any(token in text for token in ('mechanical','mech','fabrication','fabricated','sheet','metal','plastic','rubber','printed','milled')):
        return force_cdp_if_mechanical(item, {'item_type':'CDP','confidence':80,'reason':'Non-electronic mechanical part inferred from description','ambiguous':False})

    llm_result = classify_bom_line_with_llm(item)
    if llm_result:
        return force_cdp_if_mechanical(item, llm_result)

    return force_cdp_if_mechanical(item, {'item_type':'BOP','confidence':45,'reason':'Insufficient data; needs user review','ambiguous':True})


def force_cdp_if_mechanical(item, result=None):
    text = " ".join(str(item.get(k, '') or '') for k in ('MPN', 'mpn', 'part_name', 'Description', 'description', 'note')).lower()
    mechanical_tokens = [
        'custom','machined','mould','mold','enclosure','sheet metal','sheet-metal',
        'cnc','fabricated','extrusion','harness','rubber','mjf','sls','sla','fdm',
        '3d print','3d-printed','printed','standoff','spacer','cover','housing','chassis',
        'knob','shaft','gear','bushing','fixture','jig','gasket','panel','clip','latch',
        'hinge','spring','fastener','washer','nut','bolt','screw','metal','plastic',
        'insert','base','body','frame','carrier','guide','holder','mechanical','mech',
        'light pipe','top body','bottom body','faceplate','front cover','rear cover',
        'shell','bezel','trim',
    ]
    if 'button' in text and 'switch' not in text:
        return {
            'item_type': 'CDP',
            'confidence': 94,
            'reason': 'Mechanical control/button part detected',
            'ambiguous': False,
        }
    if any(token in text for token in mechanical_tokens):
        return {
            'item_type': 'CDP',
            'confidence': 97,
            'reason': 'Mechanical/custom part overridden to CDP',
            'ambiguous': False,
        }
    return result or {'item_type': 'BOP', 'confidence': 50, 'reason': 'Unchanged', 'ambiguous': True}


def classify_bom_line_with_llm(item):
    try:
        import json
        import os
        import urllib.request
        import urllib.error

        text = " ".join(str(item.get(k,'') or '') for k in ('MPN','mpn','part_name','Description','description','note')).strip()
        payload = {
            "part": {
                "mpn": str(item.get('MPN') or item.get('mpn') or item.get('part_name') or '').strip(),
                "description": str(item.get('Description') or item.get('description') or item.get('note') or '').strip(),
                "raw_text": text,
            },
            "instruction": (
                "Classify this BOM line as exactly one of: BOP, CDP, ASSEMBLY. "
                "Use CDP for mechanical/custom/fabricated parts such as enclosure, bracket, housing, knob, rubber, molded, "
                "sheet metal, CNC, MJF/SLS/SLA/3D printed, harness, fasteners, spacers, panels, covers, fixtures. "
                "Use BOP only for standard electronic purchasable components like ICs, passives, connectors, LEDs, switches, "
                "crystals, diodes, transistors. Return strict JSON with keys item_type, confidence, reason."
            )
        }

        endpoint = os.getenv('BOM_LLM_CLASSIFIER_URL', '').strip()

        def _parse_llm_text(text):
            if not text:
                return None
            try:
                parsed = json.loads(text)
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                pass
            start = text.find('{')
            end = text.rfind('}')
            if start >= 0 and end > start:
                try:
                    parsed = json.loads(text[start:end + 1])
                    if isinstance(parsed, dict):
                        return parsed
                except Exception:
                    return None
            return None

        def _normalize_result(data):
            if not isinstance(data, dict):
                return None
            item_type = str(data.get('item_type') or data.get('class') or data.get('category') or '').upper().strip()
            if item_type in {'BOP', 'CDP', 'ASSEMBLY'}:
                confidence = int(float(data.get('confidence') or 50))
                reason = str(data.get('reason') or 'LLM classification')
                return {
                    'item_type': item_type,
                    'confidence': confidence,
                    'reason': reason,
                    'ambiguous': confidence < 75,
                }
            return None

        if endpoint:
            req = urllib.request.Request(
                endpoint,
                data=json.dumps(payload).encode('utf-8'),
                headers={'Content-Type': 'application/json'},
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode('utf-8', errors='ignore')
            parsed = _normalize_result(_parse_llm_text(raw))
            if parsed:
                return parsed

        prompt_text = (
            "Return strict JSON only. "
            "Classify the following BOM line as exactly one of BOP, CDP, or ASSEMBLY. "
            "Use CDP for mechanical/custom/fabricated parts such as enclosure, bracket, housing, knob, rubber, molded, "
            "sheet metal, CNC, MJF/SLS/SLA/3D printed, harness, fasteners, spacers, panels, covers, fixtures. "
            "Use BOP only for standard electronic purchasable components like ICs, passives, connectors, LEDs, switches, "
            "crystals, diodes, transistors.\n\n"
            f"Part: {json.dumps(payload['part'], ensure_ascii=False)}"
        )

        def _openai_call(api_key):
            model = os.getenv('BOM_OPENAI_MODEL') or os.getenv('OPENAI_MODEL') or 'gpt-4o-mini'
            body = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "You classify BOM rows into BOP, CDP, or ASSEMBLY."},
                    {"role": "user", "content": prompt_text},
                ],
                "temperature": 0,
            }
            req = urllib.request.Request(
                'https://api.openai.com/v1/chat/completions',
                data=json.dumps(body).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'Authorization': f'Bearer {api_key}',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                raw = json.loads(resp.read().decode('utf-8', errors='ignore'))
            content = (((raw or {}).get('choices') or [{}])[0].get('message') or {}).get('content') or ''
            return _normalize_result(_parse_llm_text(content))

        def _anthropic_call(api_key):
            model = os.getenv('BOM_ANTHROPIC_MODEL') or 'claude-3-5-sonnet-latest'
            body = {
                "model": model,
                "max_tokens": 256,
                "temperature": 0,
                "messages": [
                    {"role": "user", "content": prompt_text},
                ],
            }
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages',
                data=json.dumps(body).encode('utf-8'),
                headers={
                    'Content-Type': 'application/json',
                    'x-api-key': api_key,
                    'anthropic-version': '2023-06-01',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                raw = json.loads(resp.read().decode('utf-8', errors='ignore'))
            content = ''.join((block.get('text') or '') for block in (raw.get('content') or []) if isinstance(block, dict))
            return _normalize_result(_parse_llm_text(content))

        def _gemini_call(api_key):
            model = os.getenv('BOM_GEMINI_MODEL') or 'gemini-1.5-flash'
            body = {
                "contents": [
                    {"parts": [{"text": prompt_text}]}
                ],
                "generationConfig": {"temperature": 0, "maxOutputTokens": 256}
            }
            url = f'https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}'
            req = urllib.request.Request(
                url,
                data=json.dumps(body).encode('utf-8'),
                headers={'Content-Type': 'application/json'},
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                raw = json.loads(resp.read().decode('utf-8', errors='ignore'))
            candidates = raw.get('candidates') or []
            content = ''
            if candidates:
                parts = (((candidates[0].get('content') or {}).get('parts')) or [])
                content = ''.join((part.get('text') or '') for part in parts if isinstance(part, dict))
            return _normalize_result(_parse_llm_text(content))

        openai_key = os.getenv('OPENAI_API_KEY', '').strip()
        anthropic_key = os.getenv('ANTHROPIC_API_KEY', '').strip()
        gemini_key = os.getenv('GEMINI_API_KEY', '').strip()

        for candidate in (
            (openai_key, _openai_call),
            (anthropic_key, _anthropic_call),
            (gemini_key, _gemini_call),
        ):
            key, caller = candidate
            if not key:
                continue
            parsed = caller(key)
            if parsed:
                return parsed
    except Exception:
        return None
    return None

def normalize_manual_bom(text, pcb_qty=1):
    rows = []
    reader = csv.reader(io.StringIO(str(text or '').replace('\t', ',')))
    for raw in reader:
        cells = [c.strip() for c in raw if c and c.strip()]
        if not cells: continue
        joined = " ".join(cells).lower()
        if 'mpn' in joined and ('qty' in joined or 'quantity' in joined): continue
        mpn = cells[0]
        qty = 1
        desc_parts = []
        for cell in cells[1:]:
            if re.fullmatch(r'\d+(\.\d+)?', cell) and qty == 1:
                qty = safe_int(cell, 1)
            else:
                desc_parts.append(cell)
        desc = " ".join(desc_parts)
        if not mpn or len(mpn) < 2: continue
        qty = max(qty * safe_int(pcb_qty, 1), 1)
        base = {'MPN': mpn, 'Description': desc, 'Quantity': qty, 'original_quantity': max(qty // max(safe_int(pcb_qty,1),1),1), 'pcb_quantity': safe_int(pcb_qty,1)}
        base.update(classify_bom_line(base))
        rows.append(base)
    return rows

def mask_email(email):
    try:
        local, domain = email.split('@')
        domain_parts = domain.split('.')
        masked_domain = domain_parts[0][:2] + '*'*max(0,len(domain_parts[0])-2)
        ext = domain_parts[-1] if len(domain_parts)>1 else 'com'
        return local[:2] + '*'*max(0,len(local)-2) + '@' + masked_domain + '.' + ext
    except: return '****@****.com'

def detect_type_col(df):
    for col in df.columns:
        if _score_col(col, TYPE_KW) >= 60:
            return col
    for col in df.columns:
        vals = {str(v).strip().upper() for v in df[col].dropna().head(50)}
        if 'CDP' in vals or 'BOP' in vals:
            return col
    return None

def row_flow(row, type_col):
    if not type_col:
        return 'BOP'
    val = str(row.get(type_col, '')).strip().upper()
    return 'CDP' if val == 'CDP' else 'BOP'

def classify_section(text):
    t = str(text or '').strip().lower()
    if not t:
        return None
    if 'cdp' in t or 'custom design part' in t:
        return 'CDP'
    if 'mbom' in t or 'mechanical' in t:
        return 'MECHANICAL'
    if 'main pcb' in t or 'display pcb' in t or 'electronics item' in t or 'ebom' in t:
        return 'BOP'
    return None

def is_section_row(values):
    non_empty = [str(v).strip() for v in values if str(v).strip() and str(v).strip().lower() not in ('nan','none')]
    if not non_empty:
        return False
    text = ' '.join(non_empty)
    if classify_section(text):
        return True
    return len(non_empty) <= 2 and not any(ch.isdigit() for ch in text)

def is_valid_mpn(v):
    v = str(v).strip()
    if not v or v.lower() in ('nan','none','nm','n/a','','tbd','tbc'): return False
    if len(v)<3 or len(v)>80: return False
    return any(c.isdigit() for c in v) and any(c.isalpha() for c in v)

def is_valid_hsn(code):
    code = str(code or '').strip()
    return bool(re.fullmatch(r'\d{8}', code))

def cybex_hsn_url(hsn_code="", description=""):
    query = str(hsn_code or description or "").strip()
    if re.fullmatch(r"\d{4,8}", query):
        return f"https://www.cybex.in/indian-custom-duty/hs-{query}"
    suffix = f"?search={quote_plus(query)}" if query else ""
    return f"https://www.cybex.in/indian-custom-duty/{suffix}"

def add_duty_preview(row):
    hsn = str(row.get("hsn_code") or "").strip()
    if not is_valid_hsn(hsn):
        row["duty_preview"] = None
        row["basic_duty_rate"] = None
        row["igst_rate"] = None
        row["sws_rate"] = None
        return row
    try:
        rates = get_duty_rates(hsn)
        row["basic_duty_rate"] = rates.get("bcd")
        row["igst_rate"] = rates.get("igst")
        row["sws_rate"] = rates.get("sws_rate")
        row["duty_hsn_description"] = rates.get("description")
        row["duty_rate_source"] = rates.get("source")
        row["duty_cybex_verified"] = bool(rates.get("cybex_verified"))
        row["duty_cybex_url"] = rates.get("cybex_url", "")
        row["duty_cybex_status"] = rates.get("cybex_status", "")
        if not row["duty_cybex_verified"]:
            row["hsn_confidence"] = "Needs verification"
            row["hsn_source"] = "HSN proposed"
            row["hsn_warning"] = "HSN not duty-ready: Cybex did not return a usable duty detail page"
            row["duty_preview"] = None
            return row
        row["hsn_confidence"] = "Cybex verified"
        row["hsn_source"] = "Cybex verified"
        row["hsn_warning"] = ""
        price = float(row.get("best_price") or 0)
        qty = safe_int(row.get("quantity", 1), 1)
        row["duty_preview"] = calculate_import_duty(price, hsn, qty) if price > 0 else None
    except Exception as e:
        row["duty_preview"] = None
        row["duty_warning"] = f"Duty preview failed: {e}"
    return row

# ── BOM Column Detection ───────────────────────────────────────
MPN_KW  = ['mpn','mfr part','manufacturer part','part number','part no','partno','part_no',
           'part_number','mfrpn','mfr_pn','manufacturer part number','component','item',
           'sku','reference','part','pn','bom item','material','component number','part code','item code']
DESC_KW = ['description','desc','component description','item description','details',
           'specification','spec','name','component name','item name','product','product name',
           'title','remarks','comment','notes','value','component value','part description']
QTY_KW  = ['quantity','qty','count','units','pcs','pieces','nos','required','req',
           'req qty','required qty','total qty','amount','number','no of pcs',
           'order qty','bom qty','usage','usage qty','demand','need']
TYPE_KW = ['type','item type','category','flow','rfq type','procurement type','section']

def _score_col(name, kws):
    n = str(name).lower().strip()
    if n in kws: return 100
    for k in kws:
        if n.startswith(k) or n.endswith(k): return 80
    for k in kws:
        if k in n: return 60
    return 0

def _like_mpn(s):
    sample = s.dropna().astype(str).head(10)
    if not len(sample): return 0.0
    pat = re.compile(r'^[A-Za-z0-9][A-Za-z0-9\-_\.\/]{2,40}$')
    return sum(1 for v in sample if pat.match(v.strip()))/len(sample)

def _like_qty(s):
    sample = s.dropna().head(10)
    if not len(sample): return 0.0
    return sum(1 for v in sample if str(v).strip().replace('.','').isdigit())/len(sample)

def _like_desc(s):
    sample = s.dropna().astype(str).head(10)
    if not len(sample): return 0.0
    return sum(1 for v in sample if len(v.strip())>5)/len(sample)

def detect_columns(df):
    cols   = list(df.columns)
    scores = {c:{'mpn':0,'desc':0,'qty':0} for c in cols}
    for c in cols:
        scores[c]['mpn']  = _score_col(c, MPN_KW)
        scores[c]['desc'] = _score_col(c, DESC_KW)
        scores[c]['qty']  = _score_col(c, QTY_KW)
    for c in cols:
        scores[c]['mpn']  += int(_like_mpn(df[c])  * 40)
        scores[c]['qty']  += int(_like_qty(df[c])  * 40)
        scores[c]['desc'] += int(_like_desc(df[c]) * 20)
    used = set()
    def best(t):
        bc, bs = None, -1
        for c in cols:
            if c in used: continue
            if scores[c][t] > bs: bs=scores[c][t]; bc=c
        if bc: used.add(bc)
        return bc, bs
    mc, ms = best('mpn')
    dc, ds = best('desc')
    qc, qs = best('qty')
    method = 'header'
    if ms < 50: method = 'data_pattern'
    if not mc or not qc:
        method = 'fallback'
        mc = cols[0] if cols else None
        dc = cols[1] if len(cols)>1 else None
        qc = cols[2] if len(cols)>2 else None
    return {'mpn_col':mc,'desc_col':dc,'qty_col':qc,'all_columns':cols,'method':method,
            'confidence':{'mpn':ms,'desc':ds,'qty':qs}}

# ── Routes ─────────────────────────────────────────────────────
def index(request):
    try:
        return render_template('index.html')
    except TemplateNotFound:
        return (
            "Template missing: templates/index.html. "
            "Create a templates folder inside BOM-Tool and put index.html inside it.",
            404,
        )

def healthz(request):
    try:
        conn = get_db()
        conn.execute("SELECT 1")
        conn.close()
        return jsonify({"ok": True, "database": "ok"}), 200
    except Exception as exc:
        return jsonify({"ok": False, "database": "error", "error": str(exc)}), 500

def runtime_diagnostics(request):
    """Non-secret deployment diagnostics."""
    def present(name):
        value = os.getenv(name, "")
        return bool(str(value).strip())

    return jsonify({
        "ok": True,
        "database_url_set": present("DATABASE_URL"),
        "providers": {
            "mouser": present("MOUSER_API_KEY"),
            "digikey": present("DIGIKEY_CLIENT_ID") and present("DIGIKEY_CLIENT_SECRET"),
            "element14": present("ELEMENT14_API_KEY"),
            "arrow": present("ARROW_API_KEY"),
            "jina": present("JINA_API_KEY"),
            "openai": present("OPENAI_API_KEY"),
            "gemini": present("GEMINI_API_KEY") or present("GOOGLE_API_KEY"),
            "anthropic": present("ANTHROPIC_API_KEY"),
        },
        "message": "No secrets are returned; `false` means the env var is unset in the Container App.",
    })

def upload_bom(request):
    if 'file' not in request.files: return jsonify({"error":"No file"}), 400
    file = request.files['file']
    if not is_spreadsheet_file(file.filename, file.mimetype):
        return jsonify({"error":"Please upload a CSV, XLS or XLSX BOM file"}), 400
    filename = secure_filename(file.filename or "bom_upload.csv")
    fp   = os.path.join(UPLOAD_FOLDER, filename)
    file.save(fp)
    pcb_qty = safe_int(request.form.get('pcb_quantity', 1), 1)
    try:
        is_csv = is_csv_file(filename, file.mimetype)
        sheet_names = [0] if is_csv else pd.ExcelFile(fp).sheet_names
        comps, cdp_comps, mechanical_comps, dnp_comps, seen = [], [], [], [], set()
        det = {}

        for sheet in sheet_names:
            df_raw = read_csv_safely(fp,header=None) if is_csv else pd.read_excel(fp,header=None,sheet_name=sheet)
            if df_raw.empty:
                continue
            sheet_name = str(sheet)
            sheet_kind = 'MECHANICAL' if sheet_name.strip().lower() == 'mbom' or 'mechanical' in sheet_name.lower() else None

            hrow, hscore = 0, 0
            for i in range(min(20,len(df_raw))):
                rv = [str(v).lower().strip() for v in df_raw.iloc[i].tolist() if str(v) not in ('nan','None','')]
                sc = sum(1 for k in ['mpn','part','description','qty','quantity','item','manufacturer','part name'] if k in ' '.join(rv))
                if sc > hscore: hscore=sc; hrow=i

            df = read_csv_safely(fp,header=hrow) if is_csv else pd.read_excel(fp,header=hrow,sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            df = df.dropna(how='all')
            if df.empty:
                continue

            local_det = detect_columns(df)
            if not det:
                det = local_det
            mc, dc, qc = local_det['mpn_col'], local_det['desc_col'], local_det['qty_col']
            tc = detect_type_col(df)
            det['type_col'] = tc
            dnp_col = None
            for col in df.columns:
                cn = str(col).strip().lower()
                if any(k in cn for k in ("dnp", "dnf", "populate", "mount", "fit status", "assembly option")):
                    dnp_col = col
                    break
            det['dnp_col'] = dnp_col
            current_section = sheet_kind or 'BOP'

            for i in range(len(df)):
                row = df.iloc[i]
                raw_vals = [clean_nan(v,'') for v in row.tolist()]
                joined = ' '.join(str(v).strip() for v in raw_vals if str(v).strip())
                sec = classify_section(joined)
                if sec:
                    current_section = sec
                    continue
                if is_section_row(raw_vals):
                    continue

                item_name = ''
                for fc in ['Item Name','item name','Part Name','part name','Component','Name']:
                    if fc in df.columns:
                        item_name = str(clean_nan(row[fc],'')).strip()
                        if item_name and item_name.lower() not in ('nan','none',''): break

                mpn = str(row[mc]).strip().split('\n')[0].split('/')[0].strip() if mc and mc in df.columns else ''
                desc = str(clean_nan(row[dc],'')).strip() if dc and dc in df.columns else ''
                if not desc or desc.lower()=='nan':
                    desc = item_name

                qraw = row[qc] if qc and qc in df.columns else 1
                orig_qty = safe_int(qraw,1)

                flow = row_flow(row, tc) if tc else current_section
                part_name = item_name or mpn or desc
                dnp_status = row[dnp_col] if dnp_col and dnp_col in df.columns else ""
                if is_dnp_component(qraw, dnp_status, desc, item_name):
                    dnp_comps.append({
                        'MPN': mpn,
                        'Description': clean_nan(desc,'')[:200],
                        'Quantity': 0,
                        'original_quantity': max(orig_qty, 0),
                        'pcb_quantity': pcb_qty,
                        'item_type': 'DNP',
                        'status': 'DNP',
                        'dnp': True,
                        'note': 'Do not populate / excluded from pricing and RFQ'
                    })
                    continue
                if str(qraw).strip().upper() in ('N/A','NA','NO','-'): continue
                if orig_qty <= 0:
                    continue

                if not part_name:
                    continue

                if flow == 'CDP':
                    cdp_comps.append({'part_name':part_name,'Description':clean_nan(desc,'')[:200],
                                     'Quantity':orig_qty*pcb_qty,'original_quantity':orig_qty,
                                     'pcb_quantity':pcb_qty,'item_type':'CDP',
                                     'note':'Custom manufacturing required'})
                    continue

                if flow == 'MECHANICAL':
                    mechanical_comps.append({'part_name':part_name,'Description':clean_nan(desc,'')[:200],
                                             'Quantity':orig_qty*pcb_qty,'original_quantity':orig_qty,
                                             'pcb_quantity':pcb_qty,'item_type':'MECHANICAL',
                                             'note':'Mechanical manufacturing required'})
                    continue

                if not is_valid_mpn(mpn) or mpn.upper() in seen: continue
                seen.add(mpn.upper())
                comps.append({'MPN':mpn,'Description':clean_nan(desc,'')[:200],
                             'Quantity':orig_qty*pcb_qty,'original_quantity':orig_qty,
                             'pcb_quantity':pcb_qty,'item_type':'BOP'})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error":str(e)}), 500
    rfq = create_rfq(filename)
    return safe_jsonify({"success":True,"rfq_code":rfq,"bom_file":filename,
                         "total_components":len(comps),"components":comps,
                         "bop_components":comps,"cdp_components":cdp_comps,
                         "mechanical_components":mechanical_comps,
                         "dnp_components":dnp_comps,
                         "total_cdp_components":len(cdp_comps),
                         "total_mechanical_components":len(mechanical_comps),
                         "total_dnp_components":len(dnp_comps),
                         "detection":det,"pcb_quantity":pcb_qty})

def fetch_prices_bulk(request):
    global _vendor_pricing_service
    body  = request.json or {}
    comps = [clean_comp(c) for c in body.get('components',[])]
    rfq   = body.get('rfq_code','')
    fast_pricing = bool(body.get('fast_pricing'))
    deep_search = bool(body.get('deep_search'))
    no_cache = bool(body.get('no_cache'))
    # ── DigiKey token: get_digikey_token() has its own TTL cache so it returns
    # instantly on warm calls. Wrap in thread only on first/cold call to avoid
    # blocking the handler for up to 25s (Retry(total=2) × timeout=8s). ──
    from Digikey_fetch import _token_cache as _dk_cache
    _dk_warm = bool(_dk_cache.get("access_token") and time.time() < _dk_cache.get("expires_at", 0))
    if _dk_warm:
        token = get_dk_token()   # instant — cache hit, no network call
    else:
        _dk_ex = ThreadPoolExecutor(max_workers=1)
        _dk_fut = _dk_ex.submit(get_dk_token)
        _dk_done, _ = _wait_futures([_dk_fut], timeout=0.25 if fast_pricing else 5)
        token = _dk_fut.result() if _dk_done else None
        _dk_ex.shutdown(wait=False, cancel_futures=True)
    if no_cache:
        with _vendor_pricing_lock:
            _vendor_pricing_service = None

    def _m(mpn):
        try: return get_mouser_price(mpn)
        except: return None
    def _d(mpn):
        try: return get_digikey_price(mpn,token) if token else None
        except: return None
    def _e(mpn):
        try: return get_element14_price(mpn)
        except: return None
    def _l(mpn):
        try: return get_lcsc_price(mpn)
        except: return None
    def _a(mpn):
        if fast_pricing and not deep_search and not os.getenv("ARROW_API_KEY"):
            return None
        try: return get_arrow_price(mpn)
        except: return None
    def _i(mpn):
        try: return get_indian_best_price(mpn)
        except: return None

    def best_price(results):
        b, bp = None, None
        for s,info in results.items():
            p = info.get('price')
            if p and (bp is None or p < bp): b=s; bp=p
        return b, bp

    def fetch_one(comp):
        mpn  = str(comp.get('MPN','')).strip()
        qty  = positive_int(comp.get('Quantity',comp.get('Qty',1)),1)
        original_qty = positive_int(comp.get('original_quantity', qty), qty)
        pcb_qty = positive_int(comp.get('pcb_quantity', 1), 1)
        desc = str(clean_nan(comp.get('Description',''),''))
        manufacturer = str(clean_nan(comp.get('Manufacturer', comp.get('manufacturer', '')), ''))
        cached = None if no_cache else _get_cached(mpn, desc)
        if cached and not deep_search:
            qty_results = reprice_results_for_quantity(cached.get('results',{}), qty)
            b, bp = best_price(qty_results)
            row = {**cached,"results":qty_results,
                   "quantity":qty,"original_quantity":original_qty,"pcb_quantity":pcb_qty,
                   "best_supplier":b,"best_price":bp,
                   "total_cost":round(bp*qty,2) if bp else None,"cached":True}
            return row if fast_pricing else add_duty_preview(row)
        raw = {}
        def _g(mpn_):
            try: return get_google_shopping_results(mpn_, desc)
            except: return None

        vendor_jobs = [(_m, "mouser"), (_d, "digikey"), (_e, "element14"), (_l, "lcsc")]
        if deep_search or not fast_pricing:
            if not fast_pricing:
                vendor_jobs.append((_i, "indian"))
            if not fast_pricing and os.getenv("ARROW_API_KEY"):
                vendor_jobs.append((_a, "arrow"))
            if not fast_pricing and os.getenv("GOOGLE_API_KEY") and os.getenv("GOOGLE_CSE_ID"):
                vendor_jobs.append((_g, "google_shopping"))
        per_vendor_timeout = 5 if fast_pricing and deep_search else (4 if fast_pricing else 7)
        ex = ThreadPoolExecutor(max_workers=len(vendor_jobs))
        try:
            fs = {ex.submit(fn, mpn): key for fn, key in vendor_jobs}
            # ── FIXED: collect ALL vendor results in ONE parallel wait ──
            # Old code: `for f in fs: f.result(timeout=5)` waited up to
            # per_vendor_timeout EACH → up to N×5s serial.
            # New code: wait for all to finish in ONE go → max 5s total.
            _done, _not_done = _wait_futures(fs.keys(), timeout=per_vendor_timeout)
            for f in _not_done:
                f.cancel()
                raw[fs[f]] = None
            for f in _done:
                try: raw[fs[f]] = f.result()
                except: raw[fs[f]] = None
        finally:
            ex.shutdown(wait=False, cancel_futures=True)
        supplemental_candidates = []
        results = {}
        for k,r in raw.items():
            if r:
                r = apply_overseas_landed_price(k, r)
                p = to_float(r.get('price'))
                if not p: continue
                # Skip non-stock items
                stock_raw = str(clean_nan(r.get('stock'), '0')).strip().lower()
                stock_num = 0
                try:
                    stock_num = int(''.join(filter(str.isdigit, stock_raw)) or '0')
                except:
                    pass
                # Accept if stock > 0 OR stock info not available (blank/unknown)
                stock_positive_words = ("in stock", "available", "instock", "yes")
                has_positive_stock_text = any(word in stock_raw for word in stock_positive_words)
                if stock_raw not in ('', 'none', 'n/a') and stock_num == 0 and not has_positive_stock_text:
                    continue  # Skip truly out-of-stock items
                qty_price = price_for_quantity(r, qty) or p
                results[k]={'price':qty_price,'base_price':p,'price_basis_qty':qty,
                            'price_breaks':r.get('price_breaks') or [],
                            **supplier_price_summary(r, qty),
                            'stock':str(clean_nan(r.get('stock'),'')),
                            'lead_time':str(clean_nan(r.get('lead_time'),'N/A')),
                            'moq':clean_nan(r.get('moq'),1),
                            'min_price_qty':clean_nan(r.get('min_price_qty'), r.get('moq', 1)),
                            'loose_available': bool(r.get('loose_available', True)),
                            'package_type': str(clean_nan(r.get('package_type'), '')),
                            'url':str(clean_nan(r.get('url'),''))}
        if not fast_pricing and (deep_search or not fast_pricing) and not results:
            supplemental_candidates = fetch_supplemental_vendor_prices(mpn, description=desc, qty=qty)
        for candidate in supplemental_candidates:
            if safe_int(candidate.get("match_confidence", 0), 0) < 60:
                continue
            legacy = normalized_vendor_to_legacy_result(candidate, qty)
            if not legacy:
                continue
            vendor_key = str(candidate.get("vendor_name") or "Vendor").strip()
            results[vendor_key] = legacy
        # ── Shopping fallback: run whenever results < 2 (not just empty) ──
        # Previously only ran when results == 0. Now also runs when we have
        # only 1 source — Google Shopping + 7 Indian stores give 99% coverage.
        _need_shopping = deep_search and (desc or mpn) and len(results) < 3
        if _need_shopping:
            try:
                mfr_hint = manufacturer if manufacturer else ""
                _shop_ex = ThreadPoolExecutor(max_workers=1)
                _shop_f  = _shop_ex.submit(
                    shopping_fallback_search, mpn, desc, mfr_hint,
                    max_results=6,           # more results → better coverage
                    include_jina=True,       # Robu + Tenettech
                )
                _done, _ = _wait_futures([_shop_f], timeout=12 if fast_pricing else 18)
                shopping_hits = _shop_f.result() if _done else []
                _shop_ex.shutdown(wait=False, cancel_futures=True)
                for hit in shopping_hits:
                    legacy = shopping_hit_to_legacy_result(hit, qty)
                    if not legacy:
                        continue
                    vendor_key = str(hit.get("vendor_name") or "ShoppingResult").strip()
                    source = hit.get("source_type", "")
                    legacy["match_type"] = "Google" if "google" in source else "Spec"
                    legacy["description_match_warning"] = (
                        "Google Shopping result — verify specs before buying"
                        if "google" in source
                        else "Spec match from online search — verify before buying"
                    )
                    if (
                        vendor_key not in results
                        or safe_int(legacy.get("match_confidence"), 0) > safe_int(results[vendor_key].get("match_confidence"), 0)
                    ):
                        results[vendor_key] = legacy
                if shopping_hits:
                    log.info("shopping_fallback added %d result(s) for %s", len(shopping_hits), mpn)
            except Exception as _se:
                print(f"Shopping fallback error for {mpn}: {_se}")

            desc_indian = None if fast_pricing else _i(desc)
            if desc_indian and to_float(desc_indian.get("price")):
                results["indian"] = {
                    "price": to_float(desc_indian.get("price")),
                    "base_price": to_float(desc_indian.get("price")),
                    "price_basis_qty": qty,
                    "price_breaks": [],
                    **supplier_price_summary({"price": desc_indian.get("price"), "price_breaks": []}, qty),
                    "stock": str(clean_nan(desc_indian.get("stock"), "N/A")),
                    "lead_time": str(clean_nan(desc_indian.get("lead_time"), "N/A")),
                    "moq": clean_nan(desc_indian.get("moq"), 1),
                    "url": str(clean_nan(desc_indian.get("url"), "")),
                    "match_confidence": 55,
                    "source_type": "description_scrape",
                }
            if not results:
                full_desc_candidates = [] if fast_pricing else fetch_supplemental_vendor_prices(desc, description=desc, qty=qty)
                for candidate in full_desc_candidates:
                    if safe_int(candidate.get("match_confidence", 0), 0) < 55:
                        continue
                    legacy = normalized_vendor_to_legacy_result(candidate, qty)
                    if not legacy:
                        continue
                    legacy["source_type"] = "description_scrape"
                    vendor_key = str(candidate.get("vendor_name") or "Vendor").strip()
                    results[vendor_key] = legacy

            if not results and desc:
                dk_desc = _d(desc)
                if dk_desc and to_float(dk_desc.get("price")):
                    dk_desc = apply_overseas_landed_price("digikey", dk_desc)
                    p = to_float(dk_desc.get("price"))
                    results["digikey"] = {
                        "price": price_for_quantity(dk_desc, qty) or p,
                        "base_price": p,
                        "price_basis_qty": qty,
                        "price_breaks": dk_desc.get("price_breaks") or [],
                        **supplier_price_summary(dk_desc, qty),
                        "stock": str(clean_nan(dk_desc.get("stock"), "N/A")),
                        "lead_time": str(clean_nan(dk_desc.get("lead_time"), "N/A")),
                        "moq": clean_nan(dk_desc.get("moq"), 1),
                        "min_price_qty": clean_nan(dk_desc.get("min_price_qty"), dk_desc.get("moq", 1)),
                        "loose_available": bool(dk_desc.get("loose_available", True)),
                        "package_type": str(clean_nan(dk_desc.get("package_type"), "")),
                        "url": str(clean_nan(dk_desc.get("url"), "")),
                        "product_title": str(clean_nan(dk_desc.get("description"), "")),
                        "match_confidence": 62,
                        "source_type": "digikey_description_search",
                    }

            # ── Shopping fallback: spec-match search on Indian stores ───────
            # Jab MPN exact match na mile, tab description se key specs (value+package)
            # nikal ke Indian stores pe seedha search karta hai.
            # Capacitor: 100nF + 0402 dono match hone chahiye title mein.
            # Resistor:  100R  + 0402 dono match hone chahiye.
            if not fast_pricing and deep_search and not results and (desc or mpn):
                try:
                    mfr_hint = ""
                    # ── FIXED: wrap in 12s timeout (same as fast_pricing branch) ──
                    _shop2_ex = ThreadPoolExecutor(max_workers=1)
                    _shop2_f  = _shop2_ex.submit(shopping_fallback_search, mpn, desc, mfr_hint)
                    _shop2_done, _ = _wait_futures([_shop2_f], timeout=12)
                    shopping_hits = _shop2_f.result() if _shop2_done else []
                    _shop2_ex.shutdown(wait=False, cancel_futures=True)
                    for hit in shopping_hits:
                        legacy = shopping_hit_to_legacy_result(hit, qty)
                        if not legacy:
                            continue
                        vendor_key = str(hit.get("vendor_name") or "ShoppingResult").strip()
                        legacy["match_type"] = "Spec"
                        legacy["description_match_warning"] = "Spec match from online search - verify before buying"
                        if (
                            vendor_key not in results
                            or safe_int(legacy.get("match_confidence"), 0) > safe_int(results[vendor_key].get("match_confidence"), 0)
                        ):
                            results[vendor_key] = legacy
                except Exception as _se:
                    print(f"Shopping fallback error for {mpn}: {_se}")

        mr,dr,er,lr,ar,ir = raw.get('mouser'),raw.get('digikey'),raw.get('element14'),raw.get('lcsc'),raw.get('arrow'),raw.get('indian')
        mfr = ""
        if mr:  mfr = extract_field(mr,'manufacturer','Manufacturer','ManufacturerName')
        if not mfr and er:  mfr = extract_mfr_desc(er.get('description',''),mpn)
        if not mfr and dr:  mfr = extract_mfr_url(dr.get('url',''))
        if not mfr and lr:  mfr = extract_field(lr,'manufacturer') or ""
        if not mfr and ar:  mfr = extract_field(ar,'manufacturer') or ""
        if not mfr:
            for vendor_info in results.values():
                mfr = extract_field(vendor_info, "manufacturer")
                if mfr:
                    break
        if not mfr:
            mfr = infer_manufacturer_local(mpn, desc)
        hsn = extract_field(mr,'hsn_code','HsnCode') if mr else ""
        hsn = hsn if is_valid_hsn(hsn) else ""
        datasheet_url = extract_datasheet_from_sources(mr, dr, er, lr, ar, ir)
        local_hsn = {}
        local_candidate = infer_hsn_local(mpn, desc, mfr)
        local_code = str(local_candidate.get("hsn_code", "")).strip()
        if is_valid_hsn(local_code) and local_code != hsn:
            hsn = local_code
            local_hsn = local_candidate
        elif hsn:
            try:
                if not get_duty_rates(hsn).get("cybex_verified"):
                    if is_valid_hsn(local_code) and get_duty_rates(local_code).get("cybex_verified"):
                        hsn = local_code
                        local_hsn = local_candidate
            except Exception:
                pass
        if not hsn:
            local_hsn = infer_hsn_local(mpn, desc, mfr)
            if is_valid_hsn(local_hsn.get("hsn_code", "")):
                hsn = str(local_hsn.get("hsn_code", "")).strip()
            else:
                local_hsn = {}
        # Normalise Element14 / Farnell / Newark aliases → canonical "element14" key
        _E14_ALIASES = {"farnell", "farnel", "farnellindia", "element14india",
                        "element14_in", "element14in", "newark", "element14com"}
        for _k in list(results.keys()):
            _kn = _k.lower().replace(" ", "").replace(".", "").replace("_", "")
            if _k != "element14" and _kn in _E14_ALIASES:
                if "element14" not in results:
                    results["element14"] = results.pop(_k)
                else:
                    results.pop(_k)

        entry = {"mpn":mpn,"description":desc,"manufacturer":mfr,"hsn_code":hsn,
                 "hsn_desc":local_hsn.get("hsn_desc","") if local_hsn else "",
                 "hsn_confidence":local_hsn.get("confidence","High") if local_hsn else ("High" if hsn else ""),
                 "hsn_source":"Local Rule Proposed" if local_hsn else ("Distributor" if hsn else ""),
                 "hsn_warning":"Proposed HSN - verify before duty" if local_hsn else ("" if hsn else "HSN not verified yet"),
                 "hsn_reference_url":cybex_hsn_url(hsn, desc),
                 "datasheet_url":datasheet_url,
                 "results":results,"alternatives":[]}
        if not no_cache:
            _set_cached(mpn, entry, desc)
        b, bp = best_price(results)
        return {**entry,"quantity":qty,"original_quantity":original_qty,"pcb_quantity":pcb_qty,
                "best_supplier":b,"best_price":bp,
                "total_cost":round(bp*qty,2) if bp else None,"cached":False}

    all_r = []
    ex = ThreadPoolExecutor(max_workers=min(max(len(comps), 1), 8))
    try:
        futures = {ex.submit(fetch_one, comp): comp for comp in comps}
        done, pending = _wait_futures(futures.keys(), timeout=14 if fast_pricing else 22)
        for fut in pending:
            fut.cancel()
            comp = futures[fut]
            mpn = str(comp.get('MPN', '')).strip()
            qty = positive_int(comp.get('Quantity', comp.get('Qty', 1)), 1)
            desc = str(clean_nan(comp.get('Description', ''), ''))
            mfr = str(clean_nan(comp.get('Manufacturer', comp.get('manufacturer', '')), ''))
            hsn_local = infer_hsn_local(mpn, desc, mfr)
            timeout_row = {
                "mpn": mpn,
                "description": desc,
                "manufacturer": mfr or infer_manufacturer_local(mpn, desc),
                "hsn_code": hsn_local.get("hsn_code", "N/A"),
                "hsn_desc": hsn_local.get("hsn_desc", ""),
                "hsn_confidence": hsn_local.get("confidence", "Low"),
                "hsn_source": "Local Rule Proposed",
                "hsn_warning": "Pricing timeout - sent to RFQ/deep review",
                "hsn_reference_url": cybex_hsn_url(hsn_local.get("hsn_code", ""), desc),
                "datasheet_url": "",
                "results": {},
                "alternatives": [],
                "quantity": qty,
                "original_quantity": positive_int(comp.get('original_quantity', qty), qty),
                "pcb_quantity": positive_int(comp.get('pcb_quantity', 1), 1),
                "best_supplier": None,
                "best_price": None,
                "total_cost": None,
                "cached": False,
                "pricing_status": "timeout_rfq",
            }
            all_r.append(timeout_row if fast_pricing else add_duty_preview(timeout_row))
        for fut in done:
            try:
                all_r.append(fut.result())
            except Exception:
                pass
    finally:
        ex.shutdown(wait=False, cancel_futures=True)

    if rfq:
        try: save_price_history(rfq,all_r)
        except Exception as e: print(f"History error: {e}")

    need_ai = [] if fast_pricing else [
        {"mpn": r["mpn"], "description": r["description"], "manufacturer": r.get("manufacturer", "")}
        for r in all_r if not r.get("manufacturer") or not r.get("hsn_code") or r.get("hsn_code") == "N/A"
    ]
    if need_ai:
        ai = get_hsn_bulk(need_ai)
        for r in all_r:
            a = ai.get(r["mpn"],{})
            if not r.get("manufacturer") and a.get("manufacturer"): r["manufacturer"]=a["manufacturer"]
            if not r.get("manufacturer"):
                r["manufacturer"] = infer_manufacturer_local(r.get("mpn",""), r.get("description",""))
            if (not r.get("hsn_code") or r.get("hsn_code")=="N/A") and a.get("hsn_code"):
                ai_hsn = str(a.get("hsn_code","")).strip()
                ai_conf = str(a.get("confidence","Low")).strip()
                if is_valid_hsn(ai_hsn):
                    r["hsn_code"]=ai_hsn; r["hsn_desc"]=a.get("hsn_desc","N/A")
                    r["hsn_confidence"]=ai_conf
                    r["hsn_source"]="AI Proposed" if ai_conf.lower() != "high" else "AI High Confidence"
                    r["hsn_warning"]="" if ai_conf.lower() == "high" else "Proposed HSN - verify before duty"
                    r["hsn_reference_url"]=cybex_hsn_url(ai_hsn, r.get("description",""))
                else:
                    r["hsn_code"]="N/A"; r["hsn_desc"]=a.get("hsn_desc","N/A")
                    r["hsn_confidence"]=ai_conf; r["hsn_source"]="OpenAI not accepted"
                    r["hsn_warning"]="HSN requires manual verification"

    alt_targets = [] if fast_pricing else [
        {"mpn": r["mpn"], "description": r["description"], "manufacturer": r.get("manufacturer", "")}
        for r in all_r
    ]
    if alt_targets:
        try:
            am = get_alternatives_bulk(alt_targets)
            for r in all_r:
                ar = am.get(r["mpn"],{})
                r["alternatives"]=ar.get("alternatives",[])
                r["alt_status"]=ar.get("status","no_suggestions")
                r["original_specs"]=ar.get("original_specs",{})
                if not r.get("datasheet_url"):
                    r["datasheet_url"]=ar.get("datasheet_url","")
        except Exception as e: print(f"Alt error: {e}")

    # EOL / Lifecycle status for all components
    eol_targets = [] if fast_pricing else [
        {"mpn": r["mpn"], "description": r.get("description", ""), "manufacturer": r.get("manufacturer", "")}
        for r in all_r if r.get("mpn")
    ]
    if eol_targets:
        try:
            eol_map = get_eol_bulk(eol_targets)
            for r in all_r:
                eol = eol_map.get(r["mpn"], {})
                r["eol_status"]      = eol.get("status","Unknown")
                r["eol_risk"]        = eol.get("risk","unknown")
                r["eol_color"]       = eol.get("color","muted")
                r["eol_replacement"] = eol.get("replacement","")
                r["eol_rohs"]        = eol.get("rohs","")
                r["eol_date"]        = eol.get("eol_date","")
                r["eol_note"]        = eol.get("note","")
                r["eol_source"]      = eol.get("source","")
                apply_lifecycle_fallback(r)
        except Exception as e: print(f"EOL error: {e}")

    # ── Post-process: MOQ / lead-time / lifecycle / duty ─────────────────────
    # FIXED: was sequential (8s Cybex call × N components = very slow).
    # Now runs in parallel so Cybex lookups happen concurrently.
    def _post_process(r):
        bs = r.get("best_supplier")
        si = r.get("results", {}).get(bs, {}) if bs else {}
        if not si:
            for v in r.get("results", {}).values():
                si = v
                break
        r["moq"]       = si.get("moq", 1)
        r["lead_time"] = si.get("lead_time", "N/A")
        apply_lifecycle_fallback(r)
        if not fast_pricing:
            add_duty_preview(r)
        return r

    if all_r:
        with ThreadPoolExecutor(max_workers=min(len(all_r), 8)) as pp_ex:
            pp_futures = [pp_ex.submit(_post_process, r) for r in all_r]
            done, pending = _wait_futures(pp_futures, timeout=8 if fast_pricing else 18)
            for fut in pending:
                fut.cancel()
            processed = []
            for fut, original in zip(pp_futures, all_r):
                if fut in done:
                    try:
                        processed.append(fut.result())
                        continue
                    except Exception:
                        pass
                processed.append(original)
            all_r = processed

    return safe_jsonify(all_r)

# ── Price History ──────────────────────────────────────────────
def enrich_bom_lines(request):
    body = request.json or {}
    rows = []
    for c in body.get("components", []) or []:
        row = dict(c)
        mpn = str(row.get("mpn") or row.get("MPN") or "").strip()
        if not mpn:
            continue
        row["mpn"] = mpn
        row["description"] = str(clean_nan(row.get("description", row.get("Description", "")), ""))
        row["manufacturer"] = str(clean_nan(row.get("manufacturer", row.get("Manufacturer", "")), ""))
        row.setdefault("results", row.get("results") or {})
        rows.append(row)

    targets = [
        {"mpn": r["mpn"], "description": r.get("description", ""), "manufacturer": r.get("manufacturer", "")}
        for r in rows
    ]
    if targets:
        try:
            alt_ex = ThreadPoolExecutor(max_workers=1)
            alt_future = alt_ex.submit(get_alternatives_bulk, targets)
            alt_done, _ = _wait_futures([alt_future], timeout=45)
            alt_map = alt_future.result() if alt_done else {}
            alt_ex.shutdown(wait=False, cancel_futures=True)
            for r in rows:
                alt = alt_map.get(r["mpn"], {})
                r["alternatives"] = alt.get("alternatives", [])
                r["alt_status"] = alt.get("status", "no_suggestions")
                r["original_specs"] = alt.get("original_specs", {})
                if not r.get("datasheet_url"):
                    r["datasheet_url"] = alt.get("datasheet_url", "")
        except Exception as exc:
            print(f"Enrich alternatives error: {exc}")

        try:
            eol_ex = ThreadPoolExecutor(max_workers=1)
            eol_future = eol_ex.submit(get_eol_bulk, targets)
            eol_done, _ = _wait_futures([eol_future], timeout=35)
            eol_map = eol_future.result() if eol_done else {}
            eol_ex.shutdown(wait=False, cancel_futures=True)
            for r in rows:
                eol = eol_map.get(r["mpn"], {})
                r["eol_status"] = eol.get("status", r.get("eol_status", "Unknown"))
                r["eol_risk"] = eol.get("risk", r.get("eol_risk", "unknown"))
                r["eol_color"] = eol.get("color", r.get("eol_color", "muted"))
                r["eol_replacement"] = eol.get("replacement", r.get("eol_replacement", ""))
                r["eol_rohs"] = eol.get("rohs", r.get("eol_rohs", ""))
                r["eol_date"] = eol.get("eol_date", r.get("eol_date", ""))
                r["eol_note"] = eol.get("note", r.get("eol_note", ""))
                r["eol_source"] = eol.get("source", r.get("eol_source", ""))
                apply_lifecycle_fallback(r)
        except Exception as exc:
            print(f"Enrich EOL error: {exc}")
            for r in rows:
                apply_lifecycle_fallback(r)

    return safe_jsonify(rows)

def price_history(request, mpn):
    data = get_price_history(mpn,supplier=request.args.get('supplier'),limit=int(request.args.get('limit',50)))
    return safe_jsonify({"mpn":mpn,"history":data,"count":len(data)})

def price_trend(request, mpn): return safe_jsonify(get_price_trend(mpn))

def cheapest_ever(request, mpn):
    r = get_cheapest_ever(mpn)
    return safe_jsonify(r or {"message":"No history"})

# ── Vendors ────────────────────────────────────────────────────
def get_vendors(request):
    vendors = get_all_vendors()
    result  = []
    for i,v in enumerate(vendors):
        result.append({"vendor_code":v['vendor_code'],"vendor_display":f"Vendor {i+1}",
                       "vendor_email_masked":mask_email(v['vendor_email']),
                       "vendor_name":v['vendor_name'],"category":v.get('category','General')})
    return jsonify(result)

def add_vendor_api(request):
    b = request.json or {}
    code = add_vendor(b.get('name'),b.get('email'),b.get('category','General'))
    return jsonify({"success":True,"vendor_code":code})

def delete_vendor_api(request, vendor_code):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM vendor_prices WHERE vendor_code=?",(vendor_code,))
        cur.execute("DELETE FROM rfq_vendors  WHERE vendor_code=?",(vendor_code,))
        cur.execute("DELETE FROM vendors       WHERE vendor_code=?",(vendor_code,))
        conn.commit(); conn.close()
        return jsonify({"success":True})
    except Exception as e: return jsonify({"success":False,"error":str(e)}), 500

def upload_vendors(request):
    if 'file' not in request.files: return jsonify({"error":"No file"}), 400
    file = request.files['file']
    fp   = os.path.join(UPLOAD_FOLDER,file.filename)
    file.save(fp)
    try:
        df = pd.read_csv(fp) if file.filename.endswith('.csv') else pd.read_excel(fp)
        df.columns = [str(c).lower().strip() for c in df.columns]
        nc = next((c for c in df.columns if 'name' in c),None) or df.columns[0]
        ec = next((c for c in df.columns if 'email' in c or 'mail' in c),None) or df.columns[1]
        cc = next((c for c in df.columns if 'cat' in c or 'type' in c or 'category' in c),None)
        added=skipped=0
        for row in df.itertuples(index=False):
            name  = str(getattr(row,nc,'')).strip()
            email = str(getattr(row,ec,'')).strip()
            cat   = str(getattr(row,cc,'General')).strip() if cc else 'General'
            if not name or not email or '@' not in email or email.lower() in ('nan','none',''):
                skipped+=1; continue
            try: add_vendor(name,email,cat); added+=1
            except: skipped+=1
        return jsonify({"success":True,"added":added,"skipped":skipped})
    except Exception as e: return jsonify({"error":str(e)}), 500

def get_contract_manufacturers(request):
    cms = get_all_contract_manufacturers()
    result = []
    for i,cm in enumerate(cms):
        result.append({"cm_code":cm['cm_code'],"cm_display":f"CM {i+1}",
                       "cm_email_masked":mask_email(cm['cm_email']),
                       "cm_name":cm['cm_name'],"category":cm.get('category','CDP')})
    return jsonify(result)

def add_contract_manufacturer_api(request):
    b = request.json or {}
    code = add_contract_manufacturer(b.get('name'),b.get('email'),b.get('category','CDP'))
    return jsonify({"success":True,"cm_code":code})

def upload_contract_manufacturers(request):
    if 'file' not in request.files: return jsonify({"error":"No file"}), 400
    file = request.files['file']
    fp   = os.path.join(UPLOAD_FOLDER,file.filename)
    file.save(fp)
    try:
        df = pd.read_csv(fp) if file.filename.endswith('.csv') else pd.read_excel(fp)
        df.columns = [str(c).lower().strip() for c in df.columns]
        nc = next((c for c in df.columns if 'name' in c or 'manufacturer' in c),None) or df.columns[0]
        ec = next((c for c in df.columns if 'email' in c or 'mail' in c),None) or df.columns[1]
        cc = next((c for c in df.columns if 'cat' in c or 'type' in c or 'category' in c),None)
        added=skipped=0
        for row in df.itertuples(index=False):
            name  = str(getattr(row,nc,'')).strip()
            email = str(getattr(row,ec,'')).strip()
            cat   = str(getattr(row,cc,'CDP')).strip() if cc else 'CDP'
            if not name or not email or '@' not in email or email.lower() in ('nan','none',''):
                skipped+=1; continue
            try: add_contract_manufacturer(name,email,cat); added+=1
            except: skipped+=1
        return jsonify({"success":True,"added":added,"skipped":skipped})
    except Exception as e: return jsonify({"error":str(e)}), 500

def delete_contract_manufacturer_api(request, cm_code):
    ok = delete_contract_manufacturer(cm_code)
    return jsonify({"success":ok})

def send_cdp_rfq_api(request):
    b = request.json or {}
    cdp_items = b.get('cdp_items',[])
    if not cdp_items:
        return jsonify({"success":False,"error":"CDP list empty"}), 400
    selected = b.get('selected_cms',[])
    all_cms = get_all_contract_manufacturers()
    cms = [cm for cm in all_cms if cm['cm_code'] in selected] if selected else all_cms
    if not cms:
        return jsonify({"success":False,"error":"Select at least one contract manufacturer"}), 400
    cdp_rfq_code = "CDP-" + create_rfq("CDP_RFQ.xlsx")
    sent = send_cdp_rfq(cdp_items, contract_manufacturers=cms, rfq_code=cdp_rfq_code)
    return jsonify({"success":True,"emails_sent":sent,"rfq_code":cdp_rfq_code})

def send_mechanical_rfq_api(request):
    b = request.json or {}
    mechanical_items = b.get('mechanical_items',[])
    if not mechanical_items:
        return jsonify({"success":False,"error":"Mechanical list empty"}), 400
    selected = b.get('selected_cms',[])
    all_cms = get_all_contract_manufacturers()
    cms = [cm for cm in all_cms if cm['cm_code'] in selected] if selected else all_cms
    if not cms:
        return jsonify({"success":False,"error":"Select at least one contract manufacturer"}), 400
    mechanical_rfq_code = "MECH-" + create_rfq("MECHANICAL_RFQ.xlsx")
    sent = send_mechanical_rfq(mechanical_items, contract_manufacturers=cms, rfq_code=mechanical_rfq_code)
    return jsonify({"success":True,"emails_sent":sent,"rfq_code":mechanical_rfq_code})

def send_rfq(request):
    b    = request.json or {}
    rfq  = b.get('rfq_code'); bf = b.get('bom_file')
    sel  = b.get('selected_vendors', [])
    missing_only = b.get('missing_only', True)  # default: only send unpriced rows

    all_v = get_all_vendors()
    vens  = [v for v in all_v if v['vendor_code'] in sel] if sel else all_v
    conn  = get_db(); cur = conn.cursor(); sent = 0

    # Build a filtered BOM Excel with only unpriced components
    bom_path = os.path.join('uploads', bf) if bf else None
    filtered_path = None
    if bom_path and os.path.exists(bom_path) and missing_only:
        try:
            import pandas as pd
            df = pd.read_excel(bom_path) if bom_path.endswith(('.xlsx', '.xls')) else pd.read_csv(bom_path)
            price_col = next(
                (c for c in df.columns if any(w in c.lower() for w in ['price', 'cost', 'rate', 'unit'])),
                None
            )
            if price_col:
                missing_df = df[pd.to_numeric(df[price_col], errors='coerce').isna()]
            else:
                missing_df = df  # no price column — send full BOM

            if missing_df.empty:
                conn.close()
                return jsonify({"success": False, "error": "All components are priced — nothing to RFQ."}), 400

            filtered_path = bom_path.replace('.xlsx', '_missing.xlsx').replace('.xls', '_missing.xlsx').replace('.csv', '_missing.xlsx')
            missing_df.to_excel(filtered_path, index=False)
        except Exception as e:
            logging.warning("Could not filter BOM: %s — sending full file", e)
            filtered_path = bom_path
    else:
        filtered_path = bom_path

    for v in vens:
        try:
            send_rfq_email(
                vendor_email=v['vendor_email'],
                vendor_name=v['vendor_name'],
                bom_file_path=filtered_path or bom_path,
                rfq_code=rfq
            )
            cur.execute(
                "INSERT OR IGNORE INTO rfq_vendors (rfq_code,vendor_code,email_sent,created_at) VALUES (?,?,1,datetime('now'))",
                (rfq, v['vendor_code'])
            )
            sent += 1
        except Exception as e:
            print(f"Email failed: {e}")

    conn.commit(); conn.close()
    return jsonify({"success": True, "emails_sent": sent, "rfq_code": rfq})

def check_replies(request):
    rfq = request.args.get('rfq_code')
    try:
        result = check_rfq_replies()
    except PermissionError:
        # Gmail not connected — return auth_required so frontend can show connect button
        from gmail_parser import get_auth_url
        redirect_uri = request.host_url.rstrip('/') + '/api/gmail-callback'
        auth_url = get_auth_url(redirect_uri)
        return jsonify({
            "success": False,
            "auth_required": True,
            "auth_url": auth_url,
            "error": "Gmail not connected. Click the link to authorize."
        }), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    conn = get_db(); cur = conn.cursor()
    if rfq:
        cur.execute("""SELECT vp.* FROM vendor_prices vp
            LEFT JOIN rfq_vendors rv ON vp.vendor_code=rv.vendor_code AND rv.rfq_code=?
            WHERE vp.rfq_code=? ORDER BY vp.vendor_code,vp.mpn""",(rfq,rfq))
    else:
        cur.execute("SELECT * FROM vendor_prices ORDER BY created_at DESC LIMIT 200")
    prices = [dict(r) for r in cur.fetchall()]; conn.close()
    vm = {}; ctr = 1
    for p in prices:
        vc = p.get('vendor_code','')
        if vc not in vm: vm[vc] = f"Vendor {ctr}"; ctr += 1
        p['vendor_display'] = vm[vc]; p.pop('vendor_code', None)
    return safe_jsonify({
        "success": True,
        "prices": prices,
        "total": len(prices),
        "emails_checked": result.get("emails_checked", 0),
        "processed": result.get("processed", 0)
    })


def gmail_auth_start(request):
    """Start Gmail OAuth — returns the auth URL."""
    try:
        from gmail_parser import get_auth_url
        redirect_uri = request.host_url.rstrip('/') + '/api/gmail-callback'
        auth_url = get_auth_url(redirect_uri)
        return jsonify({"success": True, "auth_url": auth_url})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def gmail_callback(request):
    """Handle Gmail OAuth callback — exchange code and redirect."""
    code  = request.args.get('code')
    error = request.args.get('error')
    if error:
        return f"""<html><body style='font-family:Inter,sans-serif;padding:40px;background:#f0fdf4'>
        <h2 style='color:#dc2626'>Gmail Authorization Failed</h2>
        <p>{error}</p>
        <p><a href='/'>← Back to BOM Tool</a></p></body></html>"""
    if not code:
        return """<html><body style='font-family:Inter,sans-serif;padding:40px'>
        <h2 style='color:#dc2626'>No authorization code received.</h2>
        <p><a href='/'>← Back to BOM Tool</a></p></body></html>"""
    try:
        from gmail_parser import exchange_code
        redirect_uri = request.host_url.rstrip('/') + '/api/gmail-callback'
        exchange_code(code, redirect_uri)
        return """<html><body style='font-family:Inter,sans-serif;padding:40px;background:#f0fdf4'>
        <h2 style='color:#16a34a'>✓ Gmail Connected Successfully!</h2>
        <p>Your Gmail account is now linked to BOM Tool.</p>
        <p>You can close this tab and go back to the tool.</p>
        <script>setTimeout(() => window.close(), 2000);</script>
        <p><a href='/'>← Back to BOM Tool</a></p></body></html>"""
    except Exception as e:
        return f"""<html><body style='font-family:Inter,sans-serif;padding:40px'>
        <h2 style='color:#dc2626'>Error: {e}</h2>
        <p><a href='/'>← Back to BOM Tool</a></p></body></html>"""


def gmail_status(request):
    """Check if Gmail is connected."""
    import os
    token_exists = os.path.exists("token.json")
    if token_exists:
        try:
            from gmail_parser import _get_gmail_service
            _get_gmail_service()
            return jsonify({"connected": True})
        except Exception:
            return jsonify({"connected": False})
    return jsonify({"connected": False})

def api_all_rfqs(request):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM rfq ORDER BY created_at DESC")
    rfqs = [dict(r) for r in cur.fetchall()]; conn.close()
    return jsonify(rfqs)

# ── Add Alternative ────────────────────────────────────────────
def add_alternative(request):
    b    = request.json or {}
    orig = b.get('original_mpn','').strip()
    alt  = {k:b.get(k,v) for k,v in [('mpn',''),('manufacturer',''),('description',''),
            ('pros',[]),('cons',[]),('confidence','Medium'),('datasheet_match','Manual'),('matched_specs',{})]}
    alt['mpn'] = alt['mpn'].strip()
    if not orig or not alt['mpn']: return jsonify({"success":False,"error":"MPN required"}), 400
    with _cache_lock:
        for key, e in list(_price_cache.items()):
            if not str(key).startswith(f"{orig.upper()}|"):
                continue
            data,ts = e
            data.setdefault('alternatives',[]).append(alt)
            _price_cache[key] = (data,ts)
    return jsonify({"success":True,"alternative":alt})

def prd_status(request):
    return jsonify(FEATURE_STATUS)

def vendor_pricing_health(request):
    service = get_vendor_pricing_service()
    if not service:
        return safe_jsonify({"enabled": False, "error": "vendor_pricing module not available"})
    return safe_jsonify({"enabled": True, "adapters": service.health_check()})

def manual_bom(request):
    b = request.json or {}
    rows = normalize_manual_bom(b.get('text',''), safe_int(b.get('pcb_quantity',1),1))
    return safe_jsonify({"success":True,"items":rows,"ambiguous":sum(1 for r in rows if r.get("ambiguous"))})

def classify_bom_api(request):
    b = request.json or {}
    items = b.get('items', [])
    classified = []
    for item in items:
        row = dict(item)
        row.update(classify_bom_line(row))
        classified.append(row)
    return safe_jsonify({"success":True,"items":classified,"ambiguous":sum(1 for r in classified if r.get("ambiguous"))})

def project_files_api(request):
    ensure_prd_tables()
    if request.method == 'GET':
        rfq = request.args.get('rfq_code','')
        conn = get_db(); cur = conn.cursor()
        if rfq:
            cur.execute("SELECT * FROM project_files WHERE rfq_code=? ORDER BY uploaded_at DESC",(rfq,))
        else:
            cur.execute("SELECT * FROM project_files ORDER BY uploaded_at DESC LIMIT 100")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return safe_jsonify({"success":True,"files":rows})

    files = request.files.getlist('files') or ([request.files['file']] if 'file' in request.files else [])
    rfq = request.form.get('rfq_code','')
    if not files: return jsonify({"error":"No files uploaded"}), 400
    saved = []
    conn = get_db(); cur = conn.cursor()
    for f in files:
        name = os.path.basename(f.filename or 'upload.bin')
        stream = classify_project_file(name)
        cur.execute("SELECT MAX(version) as v FROM project_files WHERE rfq_code=? AND original_name=?", (rfq,name))
        version = (cur.fetchone()['v'] or 0) + 1
        stored_name = f"{rfq or 'project'}_v{version}_{int(time.time())}_{name}"
        stored_path = os.path.join('uploads','project_files',stored_name)
        f.save(stored_path)
        size = os.path.getsize(stored_path) if os.path.exists(stored_path) else 0
        cur.execute("""INSERT INTO project_files
            (rfq_code, original_name, stored_path, stream, version, size_bytes, uploaded_at)
            VALUES (?,?,?,?,?,?,?)""", (rfq,name,stored_path,stream,version,size,now_iso()))
        saved.append({"name":name,"stream":stream,"version":version,"size_bytes":size})
    conn.commit(); conn.close()
    return safe_jsonify({"success":True,"files":saved})

def approve_hsn(request):
    ensure_prd_tables()
    b = request.json or {}
    mpn = str(b.get('mpn','')).strip()
    hsn = str(b.get('hsn_code','')).strip()
    if not mpn or not is_valid_hsn(hsn):
        return jsonify({"error":"Valid MPN and 8-digit HSN required"}), 400
    desc = str(b.get('hsn_desc','')).strip()
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT OR REPLACE INTO hsn_approvals
        (mpn, hsn_code, hsn_desc, source, approved_by, approved_at)
        VALUES (?,?,?,?,?,?)""", (mpn, hsn, desc, 'Manual Verified', b.get('approved_by','user'), now_iso()))
    conn.commit(); conn.close()
    return jsonify({"success":True,"mpn":mpn,"hsn_code":hsn,"source":"Manual Verified"})

def landed_cost_api(request):
    b = request.json or {}
    exchange = float(b.get('exchange_rate') or 1)
    freight = float(b.get('freight') or 0)
    insurance = float(b.get('insurance') or 0)
    cif_extra = freight + insurance
    comps = b.get('components', [])
    rows, total = [], {"assessable":0,"duty":0,"landed":0}
    priced = [c for c in comps if float(c.get('best_price') or 0) > 0]
    base_sum = sum(float(c.get('best_price') or 0) * int(c.get('quantity') or 1) for c in priced) or 1
    for c in priced:
        price = float(c.get('best_price') or 0) * exchange
        qty = int(c.get('quantity') or 1)
        share = (price * qty) / base_sum
        per_unit_cif_extra = (cif_extra * share / qty) if qty else 0
        d = calculate_import_duty(price + per_unit_cif_extra, str(c.get('hsn_code','N/A')), qty)
        rows.append({"mpn":c.get("mpn"),"hsn_code":c.get("hsn_code"),"duty":d})
        total["assessable"] += d.get("total_component_cost",0)
        total["duty"] += d.get("total_duty",0)
        total["landed"] += d.get("total_landed_cost",0)
    return safe_jsonify({"success":True,"components":rows,"summary":{k:round(v,2) for k,v in total.items()},"inputs":{"exchange_rate":exchange,"freight":freight,"insurance":insurance}})

def quotes_api(request):
    ensure_prd_tables()
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        b = request.json or {}
        vendor = str(b.get('vendor_name','')).strip()
        mpn = str(b.get('mpn','')).strip()
        price = float(b.get('unit_price') or 0)
        if not vendor or not mpn or price <= 0:
            conn.close(); return jsonify({"error":"Vendor, MPN and price are required"}), 400
        cur.execute("""INSERT INTO manual_quotes
            (rfq_code, stream, vendor_name, mpn, unit_price, moq, lead_time, certification, created_at)
            VALUES (?,?,?,?,?,?,?,?,?)""", (b.get('rfq_code',''), b.get('stream','BOP'), vendor, mpn, price,
            safe_int(b.get('moq',1),1), b.get('lead_time',''), b.get('certification',''), now_iso()))
        conn.commit()
    rfq = request.args.get('rfq_code','')
    if rfq:
        cur.execute("SELECT * FROM manual_quotes WHERE rfq_code=? ORDER BY mpn, unit_price",(rfq,))
    else:
        cur.execute("SELECT * FROM manual_quotes ORDER BY created_at DESC LIMIT 300")
    quotes = [dict(r) for r in cur.fetchall()]
    conn.close()
    grouped = {}
    for q in quotes:
        grouped.setdefault(q["mpn"], []).append(q)
    comparison = []
    for mpn, qs in grouped.items():
        ranked = sorted(qs, key=lambda x: (x.get("unit_price") or 10**12, safe_int(x.get("moq",1),1)))
        for i, q in enumerate(ranked[:3]):
            q["rank"] = f"L{i+1}"
        comparison.extend(ranked)
    return safe_jsonify({"success":True,"quotes":comparison})

def alternative_live_price(request):
    b = request.json or {}
    alts = b.get('alternatives', [])
    comps = [{"MPN":a.get("mpn",""),"Description":a.get("description",""),"Quantity":1} for a in alts if a.get("mpn")]
    if not comps:
        raise HTTPException(status_code=400, detail="No alternative MPNs")
    inner_req = _FlaskShim(
        method="POST",
        json_body={"components": comps, "rfq_code": b.get("rfq_code", "")},
    )
    return fetch_prices_bulk(inner_req)

def send_assembly_rfq_api(request):
    b = request.json or {}
    items = b.get('assembly_items',[])
    if not items:
        return jsonify({"success":False,"error":"Assembly/FATP list empty"}), 400
    selected = b.get('selected_cms',[])
    all_cms = get_all_contract_manufacturers()
    cms = [cm for cm in all_cms if cm['cm_code'] in selected] if selected else all_cms
    if not cms:
        return jsonify({"success":False,"error":"Select at least one contract manufacturer"}), 400
    rows = []
    for item in items:
        rows.append({"part_name": item.get("part_name") or item.get("MPN") or "Assembly",
                     "Description": item.get("Description") or item.get("description") or "Assembly/FATP scope",
                     "Quantity": item.get("Quantity") or item.get("quantity") or 1,
                     "note": "Assembly/FATP manufacturing required"})
    code = "ASM-" + create_rfq("ASSEMBLY_FATP_RFQ.xlsx")
    sent = send_mechanical_rfq(rows, contract_manufacturers=cms, rfq_code=code)
    return jsonify({"success":True,"emails_sent":sent,"rfq_code":code})

def factory_shortlist(request):
    b = request.json or {}
    required = str(b.get('required_capability','')).lower()
    qty = safe_int(b.get('quantity',1),1)
    cms = get_all_contract_manufacturers()
    shortlist = []
    for cm in cms:
        cat = str(cm.get('category','')).lower()
        score = 50
        if required and required in cat: score += 25
        if 'assembly' in required and ('assembly' in cat or 'smt' in cat): score += 20
        if qty > 1000 and any(w in cat for w in ['scale','production','smt']): score += 10
        shortlist.append({"cm_code":cm["cm_code"],"cm_display":mask_email(cm["cm_email"]),"category":cm.get("category","CDP"),"match_score":min(score,95),"badge":"Strength Recommended" if score >= 75 else "Review"})
    shortlist = sorted(shortlist, key=lambda x:x["match_score"], reverse=True)[:8]
    return safe_jsonify({"success":True,"factories":shortlist})

def negotiation_start(request):
    ensure_prd_tables()
    b = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO negotiation_rounds
        (rfq_code,target_price,round_no,status,notes,created_at) VALUES (?,?,?,?,?,?)""",
        (b.get('rfq_code',''), float(b.get('target_price') or 0), 1, 'Open',
         'Round 1 opened. Share masked L1 with L2/L3 manually or via RFQ follow-up.', now_iso()))
    conn.commit(); conn.close()
    return jsonify({"success":True,"status":"Open","max_rounds":3,"next_step":"Send counter-offer to L2/L3 vendors"})

def scheduling_request(request):
    ensure_prd_tables()
    b = request.json or {}
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO scheduling_requests
        (rfq_code,cm_name,preferred_date,attendees,status,created_at) VALUES (?,?,?,?,?,?)""",
        (b.get('rfq_code',''), b.get('cm_name',''), b.get('preferred_date',''), b.get('attendees',''),
         'Captured - calendar API integration pending', now_iso()))
    conn.commit(); conn.close()
    return jsonify({"success":True,"status":"Captured - calendar API integration pending"})

def production_tracking(request):
    ensure_prd_tables()
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        b = request.json or {}
        cur.execute("""INSERT INTO production_milestones
            (rfq_code,stage,owner,status,due_date,notes,updated_at) VALUES (?,?,?,?,?,?,?)""",
            (b.get('rfq_code',''), b.get('stage','Procurement'), b.get('owner',''), b.get('status','Pending'),
             b.get('due_date',''), b.get('notes',''), now_iso()))
        conn.commit()
    rfq = request.args.get('rfq_code','')
    if rfq:
        cur.execute("SELECT * FROM production_milestones WHERE rfq_code=? ORDER BY id",(rfq,))
    else:
        cur.execute("SELECT * FROM production_milestones ORDER BY id DESC LIMIT 100")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return safe_jsonify({"success":True,"milestones":rows})

def export_report(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rfq = request.args.get('rfq_code') or (request.json or {}).get('rfq_code','RFQ')
        rows = []
        if request.method == 'POST':
            rows = (request.json or {}).get('rows', [])
        else:
            payload = request.args.get('payload')
            rows = json.loads(payload) if payload else []
        rows = [
            r for r in rows
            if row_quantity(r, 0) > 0 and str(r.get("mpn") or r.get("MPN") or "").strip()
        ]

        # Fetch vendor prices for this RFQ
        vendor_cols = []  # list of display names
        vendor_map  = {}  # {mpn_upper: {vendor_display: price}}
        try:
            conn = get_db(); cur = conn.cursor()
            if rfq and rfq != 'RFQ':
                cur.execute("""SELECT vp.mpn, vp.unit_price, vp.lead_time, vp.vendor_code
                               FROM vendor_prices vp
                               LEFT JOIN rfq_vendors rv ON vp.vendor_code=rv.vendor_code
                               WHERE rv.rfq_code=? OR vp.rfq_code=?
                               ORDER BY vp.vendor_code,vp.mpn""",(rfq,rfq))
            else:
                cur.execute("SELECT mpn,unit_price,lead_time,vendor_code FROM vendor_prices ORDER BY created_at DESC LIMIT 500")
            vp_rows = cur.fetchall(); conn.close()
            vm = {}; ctr = 1
            for vp in vp_rows:
                vc = dict(vp)['vendor_code']
                if vc not in vm: vm[vc] = f"Vendor {ctr}"; ctr += 1
                disp = vm[vc]
                if disp not in vendor_cols: vendor_cols.append(disp)
                mpn_up = str(dict(vp)['mpn']).upper()
                vendor_map.setdefault(mpn_up, {})[disp] = dict(vp)['unit_price']
        except Exception as ve:
            print(f"Vendor price fetch error: {ve}")

        wb = openpyxl.Workbook()
        thin = Side(style="thin", color="D9E2EF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = '#,##0.00'

        def clean_cell(value):
            if value is None:
                return ""
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                return ""
            return value

        def best_total(row):
            price = to_float(row.get("best_price"))
            return round(price * row_quantity(row, 1), 2) if price else ""

        [r for r in rows if to_float(r.get("best_price"))]
        missing_rows = [r for r in rows if not to_float(r.get("best_price"))]

        # ── Sheet 1: Full Pricing Report ──────────────────────────
        ws = wb.active; ws.title = "Pricing Report"
        hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        PatternFill("solid", fgColor="FFF3CD")
        eol_fills = {
            "critical": PatternFill("solid", fgColor="FFE0E0"),
            "high":     PatternFill("solid", fgColor="FFE8CC"),
            "medium":   PatternFill("solid", fgColor="CCE5FF"),
            "low":      PatternFill("solid", fgColor="D4EDDA"),
        }

        base_headers = [
            "MPN","Description","Qty","Manufacturer","HSN Code",
            "Mouser INR","DigiKey INR","Element14 INR","LCSC INR","Arrow INR","India Store INR",
            "Best Supplier","Best Price INR","Total Cost INR",
            "Lifecycle Status","EOL Date","Suggested Replacement","RoHS",
            "BCD %","IGST %","Landed Cost INR",
            "Alt MPN 1","Alt Datasheet 1","Alt MPN 2","Alt Datasheet 2",
            "Actual Datasheet Link","Order URL (Best)",
        ] + vendor_cols

        for c, h in enumerate(base_headers, 1):
            cell = ws.cell(1, c, h)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        for r_i, r in enumerate(rows, 2):
            mpn    = r.get("mpn","")
            qty    = row_quantity(r, 1)
            alts   = r.get("alternatives", [])
            res    = r.get("results", {})
            bp     = r.get("best_supplier","")
            best_url = str(res.get(bp,{}).get("url","") if bp else "")

            # Alternatives (up to 2)
            alt1_mpn = alts[0].get("mpn","") if len(alts) > 0 else ""
            alt1_ds  = alts[0].get("datasheet_url","") if len(alts) > 0 else ""
            alt2_mpn = alts[1].get("mpn","") if len(alts) > 1 else ""
            alt2_ds  = alts[1].get("datasheet_url","") if len(alts) > 1 else ""

            # Vendor quotes
            v_prices = vendor_map.get(mpn.upper(), {})

            data = [
                mpn,
                r.get("description",""),
                qty,
                r.get("manufacturer",""),
                r.get("hsn_code",""),
                res.get("mouser",  {}).get("price",""),
                res.get("digikey", {}).get("price",""),
                res.get("element14",{}).get("price",""),
                res.get("lcsc",    {}).get("price",""),
                res.get("arrow",   {}).get("price",""),
                res.get("indian",  {}).get("price",""),
                r.get("best_supplier",""),
                to_float(r.get("best_price")) or "",
                best_total(r),
                r.get("eol_status",""),
                r.get("eol_date",""),
                r.get("eol_replacement",""),
                r.get("eol_rohs",""),
                r.get("basic_duty_rate",""),
                r.get("igst_rate",""),
                r.get("duty_preview",{}).get("total_landed_cost","") if isinstance(r.get("duty_preview"),dict) else "",
                alt1_mpn, alt1_ds, alt2_mpn, alt2_ds,
                r.get("datasheet_url",""),
                best_url,
            ] + [v_prices.get(vc,"") for vc in vendor_cols]

            for c, v in enumerate(data, 1):
                cell = ws.cell(r_i, c, clean_cell(v))
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=c in (2, 17, 22, 24))
                if c in (6,7,8,9,10,11,13,14,21) or c > 27:
                    cell.number_format = money_fmt
                # Color rows by EOL risk
                risk = r.get("eol_risk","")
                if risk in eol_fills:
                    cell.fill = eol_fills[risk]
                # Hyperlink for datasheet/URL columns
                if c in (25, 26) and v and str(v).startswith("http"):
                    cell.hyperlink = str(v)
                    cell.font = Font(color="0563C1", underline="single")

        # Auto column width
        for col_cells in ws.columns:
            length = max((len(str(cell.value or "")) for cell in col_cells), default=0)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ── Sheet 2: Not Priced — Needs RFQ ───────────────────────
        ws2 = wb.create_sheet("Needs RFQ")
        ws2_hdrs = ["MPN","Description","Qty","Manufacturer","Alt MPN 1","Alt MPN 2","Datasheet"]
        for c, h in enumerate(ws2_hdrs, 1):
            cell = ws2.cell(1, c, h); cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        ri2 = 2
        for r in missing_rows:
            if row_quantity(r, 0) > 0:
                alts = r.get("alternatives", [])
                ws2.cell(ri2,1, r.get("mpn",""))
                ws2.cell(ri2,2, r.get("description",""))
                ws2.cell(ri2,3, row_quantity(r, 1))
                ws2.cell(ri2,4, r.get("manufacturer",""))
                ws2.cell(ri2,5, alts[0].get("mpn","") if alts else "")
                ws2.cell(ri2,6, alts[1].get("mpn","") if len(alts)>1 else "")
                ds = r.get("datasheet_url","")
                c = ws2.cell(ri2,7, ds)
                if ds and ds.startswith("http"): c.hyperlink = ds; c.font = Font(color="0563C1", underline="single")
                for col in range(1, 8):
                    ws2.cell(ri2, col).border = border
                    ws2.cell(ri2, col).alignment = Alignment(vertical="top", wrap_text=True)
                ri2 += 1
        for col_cells in ws2.columns:
            length = max((len(str(cell.value or "")) for cell in col_cells), default=0)
            ws2.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 45)
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

        # ── Sheet 3: Vendor Quotes ─────────────────────────────────
        if vendor_cols:
            ws3 = wb.create_sheet("Vendor Quotes")
            v3_hdrs = ["MPN","Description","Qty"] + vendor_cols + ["Best Vendor","Best Vendor Price"]
            for c, h in enumerate(v3_hdrs, 1):
                cell = ws3.cell(1, c, h); cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
            for ri3, r in enumerate(rows, 2):
                mpn = r.get("mpn","")
                v_prices = vendor_map.get(mpn.upper(), {})
                vp_vals  = [v_prices.get(vc,"") for vc in vendor_cols]
                filled   = {k:v for k,v in v_prices.items() if v}
                bv = min(filled, key=filled.get) if filled else ""
                bvp = filled[bv] if bv else ""
                row_data = [mpn, r.get("description",""), row_quantity(r, 1)] + vp_vals + [bv, bvp]
                for c, v in enumerate(row_data, 1):
                    ws3.cell(ri3, c, clean_cell(v)).border = border
            ws3.freeze_panes = "A2"
            ws3.auto_filter.ref = ws3.dimensions

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"{rfq}_Procurement_Report.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def export_report_pdf(request):
    return """<html><body style='font-family:Arial'><h2>Procurement Report</h2><p>Use browser Print > Save as PDF.</p></body></html>"""

def vendor_quote_prices(request):
    """Return vendor reply prices for the live table."""
    rfq = request.args.get('rfq_code','')
    try:
        conn = get_db(); cur = conn.cursor()
        if rfq:
            cur.execute("""SELECT vp.mpn, vp.unit_price, vp.lead_time, vp.vendor_code, vp.created_at
                           FROM vendor_prices vp
                           LEFT JOIN rfq_vendors rv ON vp.vendor_code=rv.vendor_code
                           WHERE rv.rfq_code=? OR vp.rfq_code=?
                           ORDER BY vp.vendor_code, vp.mpn""",(rfq,rfq))
        else:
            cur.execute("SELECT mpn,unit_price,lead_time,vendor_code,created_at FROM vendor_prices ORDER BY created_at DESC LIMIT 500")
        rows = [dict(r) for r in cur.fetchall()]; conn.close()
        vm = {}; ctr = 1
        for p in rows:
            vc = p['vendor_code']
            if vc not in vm: vm[vc] = f"Vendor {ctr}"; ctr += 1
            p['vendor_display'] = vm[vc]
            del p['vendor_code']
        # Group by MPN: {mpn: [{vendor_display, price, lead_time}]}
        by_mpn = {}
        for p in rows:
            mpn = p['mpn'].upper()
            by_mpn.setdefault(mpn, []).append({
                "vendor":     p['vendor_display'],
                "price":      p['unit_price'],
                "lead_time":  p.get('lead_time','N/A'),
                "fetched_at": p.get('created_at',''),
            })
        all_vendors = sorted(set(vm.values()))
        return jsonify({"success":True,"by_mpn":by_mpn,"vendors":all_vendors})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)}), 500


def fetch_eol(request):
    """
    Async EOL endpoint — called separately from frontend after prices load.
    Uses Mouser → Gemini → OpenAI chain.
    """
    try:
        components = (request.json or {}).get('components', [])
        if not components:
            return jsonify({"success": False, "error": "No components"}), 400

        # Clean input
        comps = [{"mpn": str(c.get("mpn","")).strip(),
                  "manufacturer": str(c.get("manufacturer","")).strip(),
                  "description":  str(c.get("description","")).strip()}
                 for c in components if str(c.get("mpn","")).strip()]

        eol_results = get_eol_bulk(comps)

        # Return as {MPN_UPPER: eol_dict}
        eol_map = {mpn.upper(): data for mpn, data in eol_results.items()}
        return safe_jsonify({"success": True, "eol_map": eol_map, "count": len(eol_map)})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

# ── Custom Duty ────────────────────────────────────────────────
def custom_duty_bom(request):
    try:
        b = request.json or {}
        cleaned = []
        for c in b.get('components',[]):
            try:
                price=float(c.get('best_price') or 0); qty=int(c.get('quantity') or 1)
                hsn=str(c.get('hsn_code') or 'N/A').strip(); mpn=str(c.get('mpn') or '').strip()
                conf=str(c.get('hsn_confidence') or '').strip().lower()
                src=str(c.get('hsn_source') or '').strip().lower()
                cybex_verified = bool(c.get('duty_cybex_verified') or c.get('cybex_verified'))
                if not cybex_verified and is_valid_hsn(hsn):
                    try:
                        cybex_verified = bool(get_duty_rates(hsn).get("cybex_verified"))
                    except Exception:
                        cybex_verified = False
                verified = cybex_verified or 'manual verified' in src or 'cybex verified' in src or 'cybex verified' in conf
                if mpn and price>0 and is_valid_hsn(hsn) and verified:
                    cleaned.append({'mpn':mpn,'hsn_code':hsn,'best_price':price,'quantity':qty})
            except: continue
        if not cleaned: return jsonify({"error":"No valid priced components with verified 8-digit HSN codes"}), 400
        return safe_jsonify(calculate_bom_duties(cleaned))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error":str(e)}), 500

def auto_hsn_duty(request):
    """Auto-search HSN codes, Cybex-verify them, then calculate landed costs.

    Called when the normal duty endpoint returns no valid verified HSN codes.
    """
    try:
        from hsn_lookup import get_hsn_bulk
        b = request.json or {}
        components = b.get('components', [])

        priced = []
        for c in components:
            try:
                price = float(c.get('best_price') or 0)
                if price > 0:
                    priced.append({
                        'mpn': str(c.get('mpn') or '').strip(),
                        'description': str(c.get('description') or '').strip(),
                        'manufacturer': str(c.get('manufacturer') or '').strip(),
                        'best_price': price,
                        'quantity': int(c.get('quantity') or 1),
                        'existing_hsn': str(c.get('hsn_code') or 'N/A').strip(),
                    })
            except:
                continue

        if not priced:
            return jsonify({"error": "No priced components provided"}), 400

        hsn_results = get_hsn_bulk([
            {'mpn': c['mpn'], 'description': c['description'], 'manufacturer': c['manufacturer']}
            for c in priced
        ])

        results = []
        tc = td = tl = 0.0
        verified_count = 0

        for c in priced:
            mpn = c['mpn']
            price = c['best_price']
            qty = c['quantity']

            hsn_data = hsn_results.get(mpn, {})
            found_hsn = str(hsn_data.get('hsn_code') or 'N/A').strip()
            hsn_desc = str(hsn_data.get('hsn_desc') or '').strip()
            confidence = str(hsn_data.get('confidence') or '').strip()

            if is_valid_hsn(found_hsn):
                cybex = verify_hsn(found_hsn, c['description'], mpn)
                cybex_verified = bool(cybex.get('verified'))
                cybex_url = str(cybex.get('url') or '')
                cybex_status = str(cybex.get('status') or '')
            else:
                cybex_verified = False
                cybex_url = ''
                cybex_status = 'invalid_hsn'

            duty = None
            status = 'unverified'
            if cybex_verified and price > 0:
                duty = calculate_import_duty(price, found_hsn, qty)
                duty['mpn'] = mpn
                tc += duty.get('total_component_cost', 0)
                td += duty.get('total_duty', 0)
                tl += duty.get('total_landed_cost', 0)
                verified_count += 1
                status = 'ok'

            results.append({
                'mpn': mpn,
                'description': c['description'],
                'existing_hsn': c['existing_hsn'],
                'found_hsn': found_hsn,
                'hsn_desc': hsn_desc,
                'confidence': confidence,
                'cybex_verified': cybex_verified,
                'cybex_url': cybex_url,
                'cybex_status': cybex_status,
                'best_price': price,
                'quantity': qty,
                'status': status,
                'duty': duty,
            })

        eff = round(td / tc * 100, 2) if tc > 0 else 0
        return safe_jsonify({
            'components': results,
            'summary': {
                'total_component_cost': round(tc, 2),
                'total_duty': round(td, 2),
                'total_landed_cost': round(tl, 2),
                'effective_duty_pct': eff,
                'verified_count': verified_count,
                'total_count': len(results),
                'source': 'Auto HSN Search + Cybex Verification',
            }
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── Vendor DB (Excel-uploaded vendor master) ───────────────────────────────────

def _normalize_col(c: str) -> str:
    """Normalize a column name: lowercase, replace all non-alphanumeric with underscore, strip edges."""
    return re.sub(r'[^a-z0-9]+', '_', str(c).strip().lower()).strip('_')


def _detect_vendor_header_row(df_raw) -> int:
    """Find the row index containing the real header (scan first 6 rows for vendor-like keywords)."""
    header_kw = {'email', 'name', 'contact', 'vendor', 'category', 'phone', 'company'}
    best_row, best_score = 0, 0
    for i in range(min(6, len(df_raw))):
        row_text = ' '.join(str(v).lower() for v in df_raw.iloc[i].tolist() if str(v) not in ('nan', 'None', ''))
        score = sum(1 for kw in header_kw if kw in row_text)
        if score > best_score:
            best_score, best_row = score, i
    return best_row if best_score >= 2 else 0


def upload_vendor_excel(request):
    """Parse an Excel/CSV vendor database file and store in rfq_vendor_db."""
    try:
        import pandas as pd
        file = request.files.get('file')
        if not file:
            return jsonify({"error": "No file uploaded"}), 400
        filename = file.filename or "vendor_upload"
        replace_existing = request.form.get('replace', 'false').lower() == 'true'

        try:
            file_bytes = file.read()
            if filename.lower().endswith('.csv'):
                df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding='utf-8', errors='replace')
            else:
                raw = pd.read_excel(io.BytesIO(file_bytes), dtype=str, header=None)
                hrow = _detect_vendor_header_row(raw)
                df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, header=hrow)
        except Exception as exc:
            return jsonify({"error": f"Could not parse file: {exc}"}), 400

        # Normalize column names: strip + lowercase + replace non-alphanumeric with _
        df.columns = [_normalize_col(c) for c in df.columns]
        df = df.where(pd.notna(df), None)

        # Column name aliases → canonical field
        # "Vendor / Company Name" normalizes to "vendor_company_name"
        COL_MAP = {
            'vendor_name':    ['vendor_name','vendor_company_name','vendor','supplier','company','name','company_name'],
            'vendor_email':   ['vendor_email','email','contact_email','email_id','email_address'],
            'contact_person': ['contact_person','contact','contact_name','person'],
            'phone':          ['phone','mobile','contact_number','tel'],
            'category':       ['category','type','component_type','segment'],
            'city':           ['city','city_location','location','city_l'],
            'website':        ['website','web','url','site'],
            'mpn_keywords':   ['mpn_keywords','mpn','part_number','part_no','part_numbers','mpns'],
            'moq':            ['moq','min_order_qty','minimum_order','min_order'],
            'lead_time_days': ['lead_time_days','lead_time','delivery_days','delivery_time'],
            'unit_price':     ['unit_price','price','rate','cost'],
            'currency':       ['currency','curr'],
            'notes':          ['notes','remarks','comments','rfq_notes','brief_management_note'],
        }

        def _col(df_cols, aliases):
            for alias in aliases:
                if alias in df_cols:
                    return alias
            return None

        dcols = list(df.columns)
        field_map = {field: _col(dcols, aliases) for field, aliases in COL_MAP.items()}

        rows = []
        for _, r in df.iterrows():
            vname  = str(r[field_map['vendor_name']] or '').strip() if field_map['vendor_name'] else ''
            vemail = str(r[field_map['vendor_email']] or '').strip() if field_map['vendor_email'] else ''
            if not vname or not vemail or vname.lower() in ('nan', 'none', '') or '@' not in vemail:
                continue
            cat = str(r[field_map['category']] or 'General').strip() if field_map['category'] else 'General'
            if cat.lower() in ('nan', 'none', ''): cat = 'General'
            row = {
                'vendor_name':    vname,
                'vendor_email':   vemail,
                'contact_person': str(r[field_map['contact_person']] or '').strip() if field_map['contact_person'] else '',
                'phone':          str(r[field_map['phone']] or '').strip() if field_map['phone'] else '',
                'category':       cat,
                'rfq_type':       category_to_rfq_type(cat),
                'city':           str(r[field_map['city']] or '').strip() if field_map['city'] else '',
                'website':        str(r[field_map['website']] or '').strip() if field_map['website'] else '',
                'mpn_keywords':   str(r[field_map['mpn_keywords']] or '').strip() if field_map['mpn_keywords'] else '',
                'notes':          str(r[field_map['notes']] or '').strip() if field_map['notes'] else '',
                'currency':       str(r[field_map['currency']] or 'INR').strip() if field_map['currency'] else 'INR',
            }
            try: row['moq'] = int(float(r[field_map['moq']] or 1)) if field_map['moq'] else 1
            except: row['moq'] = 1
            try: row['lead_time_days'] = int(float(r[field_map['lead_time_days']] or 0)) if field_map['lead_time_days'] else 0
            except: row['lead_time_days'] = 0
            try: row['unit_price'] = float(r[field_map['unit_price']] or 0) if field_map['unit_price'] else None
            except: row['unit_price'] = None
            rows.append(row)

        if not rows:
            return jsonify({"error": "No valid vendor rows found. Ensure file has: vendor name + email columns"}), 400

        if replace_existing:
            clear_rfq_vendor_db()
        count = upsert_rfq_vendor_db(rows, source_file=filename)

        # Auto-match uploaded vendors against current BOM items
        vendor_matches = {}
        matched_bom_count = 0
        try:
            db = get_db()
            bom_rows = db.execute(
                "SELECT mpn, item_type, description FROM bom_items WHERE mpn IS NOT NULL AND mpn != '' LIMIT 200"
            ).fetchall()
            if bom_rows:
                components = [{"mpn": r[0], "item_type": r[1] or "BOP", "description": r[2] or ""} for r in bom_rows]
                vendor_matches = match_vendors_to_bom(components)
                matched_bom_count = len(vendor_matches)
        except Exception:
            pass

        return jsonify({
            "success": True,
            "inserted": count,
            "filename": filename,
            "rfq_type_breakdown": _rfq_type_breakdown(rows),
            "vendor_matches": vendor_matches,
            "matched_bom_items": matched_bom_count,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _rfq_type_breakdown(rows: list[dict]) -> dict:
    from collections import Counter
    c = Counter(r.get("rfq_type", "SOURCING") for r in rows)
    return dict(c)


def get_rfq_vendor_db_route(request):
    """Return all rfq_vendor_db rows with masked emails (safe for frontend)."""
    try:
        return safe_jsonify({"vendors": get_rfq_vendor_db_masked()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def clear_rfq_vendor_db_route(request):
    try:
        clear_rfq_vendor_db()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def import_default_vendor_excel(request):
    """Import vendors from the known local Vendor_Directory Excel file.

    Reads the file path from the request body (field: file_path) or uses a
    default path. This endpoint is for local use only.
    """
    try:
        b = request.json or {}
        file_path = str(b.get("file_path", "") or "").strip()
        if not file_path:
            # Try common download locations
            import pathlib
            candidates = [
                pathlib.Path.home() / "Downloads" / "Vendor_Directory_2026_v2 (1).xlsx",
                pathlib.Path.home() / "Downloads" / "Vendor_Directory_2026_v2.xlsx",
                pathlib.Path(BASE_DIR) / "vendor_directory.xlsx",
            ]
            for c in candidates:
                if c.exists():
                    file_path = str(c)
                    break
        if not file_path:
            return jsonify({"error": "No file_path provided and default path not found. "
                                     "Pass {\"file_path\": \"C:/path/to/vendor.xlsx\"} in the request body."}), 400

        import pandas as pd, pathlib
        fp = pathlib.Path(file_path)
        if not fp.exists():
            return jsonify({"error": f"File not found: {file_path}"}), 404

        replace_existing = bool(b.get("replace", True))

        # Read raw to find header row
        raw = pd.read_excel(str(fp), dtype=str, header=None)
        hrow = _detect_vendor_header_row(raw)
        df = pd.read_excel(str(fp), dtype=str, header=hrow)
        df.columns = [_normalize_col(c) for c in df.columns]
        df = df.where(pd.notna(df), None)

        # Column mapping for Vendor Directory format
        COL_MAP = {
            'vendor_name':    ['vendor_company_name', 'vendor_name', 'company', 'name', 'vendor', 'supplier'],
            'vendor_email':   ['email', 'vendor_email', 'contact_email', 'email_id'],
            'contact_person': ['contact_person', 'contact', 'person'],
            'phone':          ['phone', 'mobile', 'tel', 'contact_number'],
            'category':       ['category', 'type', 'segment'],
            'city':           ['city_location', 'city', 'location'],
            'website':        ['website', 'web', 'url', 'site'],
            'notes':          ['brief_management_note', 'notes', 'remarks', 'comments'],
        }
        dcols = list(df.columns)
        field_map = {f: next((a for a in als if a in dcols), None) for f, als in COL_MAP.items()}

        rows = []
        for _, r in df.iterrows():
            vname  = str(r[field_map['vendor_name']] or '').strip() if field_map['vendor_name'] else ''
            vemail = str(r[field_map['vendor_email']] or '').strip() if field_map['vendor_email'] else ''
            if not vname or not vemail or vname.lower() in ('nan', 'none', '') or '@' not in vemail:
                continue
            cat = str(r[field_map['category']] or 'General').strip() if field_map['category'] else 'General'
            if cat.lower() in ('nan', 'none', ''): cat = 'General'
            rows.append({
                'vendor_name':    vname,
                'vendor_email':   vemail,
                'contact_person': str(r[field_map['contact_person']] or '').strip() if field_map['contact_person'] else '',
                'phone':          str(r[field_map['phone']] or '').strip() if field_map['phone'] else '',
                'category':       cat,
                'rfq_type':       category_to_rfq_type(cat),
                'city':           str(r[field_map['city']] or '').strip() if field_map['city'] else '',
                'website':        str(r[field_map['website']] or '').strip() if field_map['website'] else '',
                'notes':          str(r[field_map['notes']] or '').strip() if field_map['notes'] else '',
                'moq':            1,
                'lead_time_days': 0,
                'currency':       'INR',
            })

        if not rows:
            return jsonify({"error": "No valid vendors found in the file. Check column format."}), 400

        if replace_existing:
            clear_rfq_vendor_db()
        count = upsert_rfq_vendor_db(rows, source_file=fp.name)
        breakdown = _rfq_type_breakdown(rows)
        return jsonify({"success": True, "inserted": count, "filename": fp.name,
                        "rfq_type_breakdown": breakdown,
                        "message": f"Imported {count} vendors from {fp.name}"})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def suggest_rfq_vendors(request):
    """Return all vendors grouped by RFQ type (SOURCING/MANUFACTURER/ASSEMBLY/TECHNOLOGY).

    Used by the dashboard RFQ Suggestion Panel to show which vendors to contact
    for each procurement need (sourcing components, EMS manufacturing, assembly).
    Emails are always masked.
    """
    try:
        grouped = get_vendors_by_rfq_type()
        totals = {t: len(v) for t, v in grouped.items()}
        return safe_jsonify({"by_type": grouped, "totals": totals, "total": sum(totals.values())})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def match_vendors_bom_route(request):
    """Match rfq_vendor_db vendors to BOM components. Returns {mpn: [vendor_info]}."""
    try:
        b = request.json or {}
        components = b.get('components', [])
        if not components:
            return jsonify({"error": "No components provided"}), 400
        matches = match_vendors_to_bom(components)
        return safe_jsonify({"matches": matches, "matched_count": len(matches)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def send_bulk_rfq_auto(request):
    """Send RFQ emails to vendors from rfq_vendor_db by their IDs (server-side only)."""
    try:
        from rfq_sender import send_email_direct
        from pathlib import Path
        import tempfile

        b = request.json or {}
        vendor_ids = [int(x) for x in b.get('vendor_ids', [])]
        components = b.get('components', [])
        rfq_code   = str(b.get('rfq_code') or f"RFQ-AUTO-{int(time.time())}")

        if not vendor_ids:
            return jsonify({"error": "No vendor IDs provided"}), 400

        # Build a simple CSV BOM for attachment
        lines = ["MPN,Description,Quantity"]
        for c in components:
            mpn  = str(c.get('mpn') or '').replace(',', ' ')
            desc = str(c.get('description') or '').replace(',', ' ')
            qty  = str(c.get('quantity') or 1)
            lines.append(f"{mpn},{desc},{qty}")
        bom_csv = "\n".join(lines).encode('utf-8')

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        tmp.write(bom_csv)
        tmp.close()
        bom_path = Path(tmp.name)

        sent = failed = 0
        results = []
        try:
            for vid in vendor_ids:
                vendor = get_rfq_vendor_actual(vid)
                if not vendor:
                    results.append({'id': vid, 'status': 'not_found'})
                    continue
                ok = send_email_direct(
                    rfq_code,
                    vendor['vendor_name'],
                    vendor['vendor_email'],
                    bom_path,
                )
                if ok:
                    sent += 1
                    results.append({'id': vid, 'vendor_name': vendor['vendor_name'], 'status': 'sent'})
                else:
                    failed += 1
                    results.append({'id': vid, 'vendor_name': vendor['vendor_name'], 'status': 'failed'})
        finally:
            bom_path.unlink(missing_ok=True)

        return jsonify({'rfq_code': rfq_code, 'sent': sent, 'failed': failed, 'results': results})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def custom_duty_component(request):
    try:
        b = request.json or {}
        return safe_jsonify(calculate_import_duty(float(b.get('price',0)),str(b.get('hsn_code','N/A')),int(b.get('quantity',1))))
    except Exception as e: return jsonify({"error":str(e)}), 500

# ── LLM-powered Indian vendor discovery (CDP + Mechanical) ────────────────────
def find_indian_vendors_llm(request):
    """
    Find verified Indian vendors for CDP/Mechanical components using LLM + web search.

    Request body:
      {
        "components": [{ "part_name": "...", "description": "...", "item_type": "CDP"|"MECHANICAL" }],
        "scale": "proto" | "sample" | "production",   (default: "sample")
        "rfq_code": "optional"
      }
    """
    try:
        from vendor_finder import find_vendors_bulk
    except ImportError as exc:
        return jsonify({"success": False, "error": f"vendor_finder module missing: {exc}"}), 500

    body = request.json or {}
    components = body.get("components", [])
    scale = str(body.get("scale", "sample") or "sample").lower().strip()
    rfq_code = str(body.get("rfq_code", "") or "").strip()
    if not components:
        return jsonify({"success": False, "error": "No components provided"}), 400

    # Deduplicate by part_name+item_type+scale
    seen: set[str] = set()
    unique: list[dict] = []
    for c in components:
        key = f"{str(c.get('part_name','')).strip().lower()}|{str(c.get('item_type','CDP')).upper()}|{scale}"
        if key not in seen:
            seen.add(key)
            c_copy = dict(c)
            c_copy["scale"] = scale
            unique.append(c_copy)

    try:
        results = find_vendors_bulk(unique, max_workers=4, scale=scale)

        # ── Save to database ─────────────────────────────────────────────────
        ensure_prd_tables()
        try:
            conn = get_db(); cur = conn.cursor()
            for item in results:
                part_name = str(item.get("part_name", "")).strip()
                item_type = str(item.get("item_type", "CDP")).strip()
                item_scale = str(item.get("scale", scale)).strip()
                for v in (item.get("vendors") or []):
                    cur.execute("""
                        INSERT OR REPLACE INTO vendor_search_results
                          (rfq_code, part_name, item_type, scale, vendor_name, website,
                           email, phone, location, supplies, relevance, why_trusted,
                           min_order, website_reachable, created_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        rfq_code, part_name, item_type, item_scale,
                        str(v.get("vendor_name", "")),
                        str(v.get("website", "")),
                        str(v.get("email", "")),
                        str(v.get("phone", "")),
                        str(v.get("location", "India")),
                        str(v.get("supplies", "")),
                        int(v.get("relevance", 0)),
                        str(v.get("why_trusted", "")),
                        str(v.get("min_order", "")),
                        1 if v.get("website_reachable") else 0,
                        now_iso(),
                    ))
            conn.commit(); conn.close()
        except Exception as db_exc:
            print(f"[find-indian-vendors-llm] DB save error: {db_exc}")

        return safe_jsonify({"success": True, "results": results, "count": len(results), "scale": scale})
    except Exception as exc:
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(exc)}), 500


def get_vendor_search_results(request):
    """
    Load previously saved vendor search results from DB.
    Query params: rfq_code (optional), part_name (optional), scale (optional)
    Returns: { results: [{part_name, item_type, scale, vendors: [...]}] }
    """
    ensure_prd_tables()
    rfq_code = request.args.get("rfq_code", "")
    part_name = request.args.get("part_name", "")
    scale = request.args.get("scale", "")
    try:
        conn = get_db(); cur = conn.cursor()
        wheres = []
        params = []
        if rfq_code:
            wheres.append("rfq_code=?"); params.append(rfq_code)
        if part_name:
            wheres.append("part_name=?"); params.append(part_name)
        if scale:
            wheres.append("scale=?"); params.append(scale)
        where_clause = ("WHERE " + " AND ".join(wheres)) if wheres else ""
        cur.execute(f"""
            SELECT part_name, item_type, scale, vendor_name, website, email, phone,
                   location, supplies, relevance, why_trusted, min_order, website_reachable
            FROM vendor_search_results
            {where_clause}
            ORDER BY part_name, relevance DESC
        """, params)
        rows = cur.fetchall()
        conn.close()

        # Group by (part_name, item_type, scale)
        grouped: dict[str, dict] = {}
        for r in rows:
            gkey = f"{r[0]}|{r[1]}|{r[2]}"
            if gkey not in grouped:
                grouped[gkey] = {"part_name": r[0], "item_type": r[1], "scale": r[2], "vendors": []}
            grouped[gkey]["vendors"].append({
                "vendor_name": r[3], "website": r[4], "email": r[5],
                "phone": r[6], "location": r[7], "supplies": r[8],
                "relevance": r[9], "why_trusted": r[10], "min_order": r[11],
                "website_reachable": bool(r[12]),
            })
        results = list(grouped.values())
        return safe_jsonify({"success": True, "results": results, "count": len(results)})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


# ── LangChain Chat Agent ──────────────────────────────────────────────────────

def chat_api(request):
    """
    Conversational BOM assistant powered by a LangChain ReAct agent.

    Request:  { "message": "What is the price of BC547?" }
    Response: { "reply": "...", "success": true }
    """
    try:
        from chat_agent import chat as agent_chat
    except ImportError as exc:
        return jsonify({"success": False, "error": f"Chat agent unavailable: {exc}. Install langchain deps."}), 503

    body = request.json or {}
    message = str(body.get("message", "")).strip()
    if not message:
        return jsonify({"success": False, "error": "message is required"}), 400

    try:
        reply = agent_chat(message)
        return safe_jsonify({"success": True, "reply": reply})
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 503
    except Exception as exc:
        logging.exception("chat_agent_error")
        return jsonify({"success": False, "error": str(exc)}), 500


# ── WhatsApp Bot (Twilio) ─────────────────────────────────────────────────────

def whatsapp_webhook_verify(request):
    """GET handler — lets browsers and Twilio verify the endpoint is reachable."""
    return "BOM Tool WhatsApp webhook is active.", 200


def whatsapp_webhook(request):
    """
    Twilio WhatsApp webhook.
    Twilio POST fields: Body, From (whatsapp:+91...), To, MessageSid, etc.
    Returns TwiML XML so Twilio delivers the reply back to the sender.

    Configure in Twilio console:
      Messaging → WhatsApp Senders → your number → Webhook URL:
      https://<your-domain>/api/whatsapp/webhook
    """
    from whatsapp_bot import handle_message, twiml_response, validate_twilio_request

    if not validate_twilio_request(request):
        logging.warning("whatsapp_webhook: invalid Twilio signature")
        return "Forbidden", 403

    sender = request.form.get("From", "").strip()
    body   = request.form.get("Body", "").strip()

    if not sender:
        return twiml_response("Could not identify sender."), 200, {"Content-Type": "text/xml"}

    logging.info("whatsapp_inbound from=%s body=%.80s", sender, body)

    reply = handle_message(sender, body)

    logging.info("whatsapp_reply to=%s reply=%.80s", sender, reply)
    return twiml_response(reply), 200, {"Content-Type": "text/xml"}


def whatsapp_status(request):
    """Twilio delivery status callback — just acknowledge it."""
    sid    = request.form.get("MessageSid", "")
    status = request.form.get("MessageStatus", "")
    logging.info("whatsapp_status sid=%s status=%s", sid, status)
    return "", 204


# ── Meta Cloud API WhatsApp ───────────────────────────────────────────────────

def meta_whatsapp_verify(request):
    """
    Meta webhook verification challenge.
    Meta sends GET with hub.mode, hub.verify_token, hub.challenge.
    Must return hub.challenge as plain text with 200 OK.

    Configure in Meta Developer console:
      App → WhatsApp → Configuration → Webhook URL:
      https://<domain>/api/whatsapp/meta/webhook
      Verify Token: <META_VERIFY_TOKEN from .env>
    """
    from whatsapp_bot import meta_verify_webhook
    mode      = request.args.get("hub.mode", "")
    token     = request.args.get("hub.verify_token", "")
    challenge = request.args.get("hub.challenge", "")
    response_body, status_code = meta_verify_webhook(mode, token, challenge)
    return response_body, status_code


def meta_whatsapp_webhook(request):
    """
    Meta incoming message webhook.
    Meta POSTs JSON payload when a user sends a WhatsApp message.
    We parse it, call the Groq agent, and reply via Meta Graph API.
    """
    from whatsapp_bot import meta_handle_incoming
    payload = request.get_json(silent=True) or {}
    logging.info("meta_whatsapp_inbound object=%s", payload.get("object", ""))
    threading.Thread(target=meta_handle_incoming, args=(payload,), daemon=True).start()
    return "OK", 200


def internal_chat(request):
    """
    Internal endpoint called by Baileys WhatsApp bot.
    Body: { "message": str, "sender": str (phone number) }
    Returns: { "reply": str }
    """
    data    = request.get_json(silent=True) or {}
    message = (data.get("message") or "").strip()
    sender  = (data.get("sender") or "unknown").strip()
    if not message:
        return jsonify({"reply": "Empty message received."}), 400
    try:
        from whatsapp_bot import handle_message
        reply = handle_message(sender, message)
        return jsonify({"reply": reply})
    except Exception:
        logging.exception("internal_chat error sender=%s", sender)
        return jsonify({"reply": "Agent error. Please try again."}), 500


def whatsapp_info(request):
    """Return WhatsApp bot status and connection info for the dashboard."""
    meta_token   = bool(os.getenv("META_WHATSAPP_TOKEN", "").strip())
    meta_phone   = bool(os.getenv("META_PHONE_NUMBER_ID", "").strip())
    twilio_sid   = bool(os.getenv("TWILIO_ACCOUNT_SID", "").strip())
    twilio_token = bool(os.getenv("TWILIO_AUTH_TOKEN", "").strip())

    channels = []
    if twilio_sid and twilio_token:
        channels.append({
            "provider": "Twilio Sandbox",
            "number": "+1 415 523 8886",
            "status": "active",
            "instructions": "Save this number on WhatsApp and send 'join no-package' to activate",
            "wa_link": "https://wa.me/14155238886?text=join%20no-package",
        })
    if meta_token and meta_phone:
        channels.append({
            "provider": "Meta Cloud API",
            "number": "+1 555 645 5477 (test)",
            "status": "active",
            "instructions": "Send any message to this number on WhatsApp to chat with BOM Tool AI",
            "wa_link": "https://wa.me/15556455477",
        })

    return jsonify({
        "enabled": len(channels) > 0,
        "channels": channels,
        "capabilities": [
            "Component price lookup (Mouser, Element14, Digikey, LCSC, Arrow)",
            "Indian store search (Robu, Evelta, ElectronicsComp, Flyrobo, etc.)",
            "Part identification and classification",
            "BOM summary and cost queries",
            "Vendor directory search",
            "Google Shopping search",
        ],
        "sample_queries": [
            "What is the price of ESP32?",
            "Find me 10K resistor 0402 on Indian stores",
            "What type of component is AO3415E?",
            "Show my BOM summary",
            "Find SOURCING vendors for PCB components",
        ],
    })


if __name__ == '__main__':
    print(
        "app_helpers.py is no longer a runnable entry — boot the FastAPI app via\n"
        "    uvicorn main:app --reload --port 8000\n"
        "Flask has been removed (Phase 3.1 of the refactor).",
        flush=True,
    )
_ORIGINAL_FORCE_CDP_IF_MECHANICAL = force_cdp_if_mechanical


def _is_standard_electronics_text(text: str) -> bool:
    normalized = (text or "").lower()
    standard_terms = (
        "resistor",
        "capacitor",
        "inductor",
        "connector",
        "switch",
        "led",
        "ic",
        "microcontroller",
        "memory",
        "sensor",
        "diode",
        "regulator",
        "transistor",
        "crystal",
        "oscillator",
        "fuse",
        "bead",
        "filter",
        "potentiometer",
        "encoder",
        "microphone",
        "usb",
        "jack",
        "socket",
        "receptacle",
    )
    return any(term in normalized for term in standard_terms)


def force_cdp_if_mechanical(text: str, current_label: str) -> Dict[str, Any]:
    if _is_standard_electronics_text(text):
        return {
            "category": "BOP",
            "confidence": 0.98,
            "reason": "Standard electronic component detected by keyword whitelist.",
        }
    return _ORIGINAL_FORCE_CDP_IF_MECHANICAL(text, current_label)
