"""
Email Sender — RFQ Dispatcher
==============================
Sends RFQ email with CLEAN BOM Excel (MPN, Description, Quantity only).
Original uploaded BOM is NOT sent — fresh clean file is generated.

Environment variables (Azure Communication Services — preferred):
    AZURE_COMM_CONNECTION_STRING — ACS connection string
    ACS_SENDER_EMAIL             — sender address (default: business@strenth.ai)

Fallback (Gmail SMTP):
    EMAIL_ADDRESS  — sender Gmail
    EMAIL_PASSWORD — Gmail App Password
"""

from __future__ import annotations
import base64, io, logging, os, smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from database import get_all_contract_manufacturers

load_dotenv()
log = logging.getLogger(__name__)

# ── Azure Communication Services config ──────────────────────────────────────
_ACS_CONN  = os.getenv("AZURE_COMM_CONNECTION_STRING", "")
_ACS_FROM  = os.getenv("ACS_SENDER_EMAIL", "business@strenth.ai")

# ── Gmail SMTP fallback ───────────────────────────────────────────────────────
_SENDER    = os.getenv("EMAIL_ADDRESS",  "")
_PASSWORD  = os.getenv("EMAIL_PASSWORD", "").replace(" ", "")
_SMTP_HOST = "smtp.gmail.com"
_SMTP_PORT = 587


def _send_via_acs(to_email: str, to_name: str, subject: str, plain: str, html: str,
                  attachment_bytes: bytes = None, attachment_name: str = None) -> bool:
    """Send email via Azure Communication Services."""
    try:
        from azure.communication.email import EmailClient
        client = EmailClient.from_connection_string(_ACS_CONN)

        message = {
            "senderAddress": _ACS_FROM,
            "recipients": {"to": [{"address": to_email, "displayName": to_name}]},
            "content": {"subject": subject, "plainText": plain, "html": html},
        }

        if attachment_bytes and attachment_name:
            message["attachments"] = [{
                "name": attachment_name,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentInBase64": base64.b64encode(attachment_bytes).decode("utf-8"),
            }]

        poller = client.begin_send(message)
        result = poller.result()
        log.info("ACS email sent -> %s | id: %s", to_email, result.get("id"))
        return True
    except Exception as exc:
        log.error("ACS send failed -> %s: %s", to_email, exc)
        return False

DNP_MARKERS = (
    "DNP", "DNF", "DO NOT POPULATE", "DO NOT FIT", "DO NOT PLACE",
    "NO MOUNT", "NO STUFF", "NOT POPULATED", "NOT FITTED", "NOT FIT",
    "NP", "NM"
)

def _is_dnp_row(*values) -> bool:
    """Detect do-not-populate BOM rows so RFQ sheets never ask vendors to quote them."""
    import re
    text = " ".join(str(v).upper() for v in values if str(v).strip() and str(v).lower() not in ("nan", "none"))
    if not text:
        return False
    return any(re.search(rf"(?<![A-Z0-9]){re.escape(marker)}(?![A-Z0-9])", text) for marker in DNP_MARKERS)


def _generate_clean_bom(bom_file_path: str, rfq_code: str) -> bytes:
    """Read original BOM, extract MPN+Description+Qty, return clean Excel bytes."""
    rows = []
    try:
        import pandas as pd
        fp = str(bom_file_path)

        # Read raw
        df_raw = pd.read_csv(fp, header=None) if fp.endswith('.csv') else pd.read_excel(fp, header=None, sheet_name=0)

        # Detect header row
        hrow, hscore = 0, 0
        for i in range(min(15, len(df_raw))):
            rv = [str(v).lower() for v in df_raw.iloc[i].tolist() if str(v) not in ('nan','None','')]
            sc = sum(1 for k in ['mpn','part','description','qty','quantity'] if k in ' '.join(rv))
            if sc > hscore: hscore = sc; hrow = i

        df = pd.read_csv(fp, header=hrow) if fp.endswith('.csv') else pd.read_excel(fp, header=hrow, sheet_name=0)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')

        def find_col(df, kws):
            for col in df.columns:
                if any(k in col.lower() for k in kws): return col
            return None

        mc = find_col(df, ['mpn','part number','part no','manufacturer part','mfr part','part'])
        dc = find_col(df, ['description','desc','specification','spec'])
        qc = find_col(df, ['quantity','qty','count','pcs'])
        tc = find_col(df, ['type','item type','category','flow','rfq type','procurement type'])

        current_section = "BOP"
        seen = set()
        for i in range(len(df)):
            row = df.iloc[i]
            row_text = " ".join(str(clean).strip().lower() for clean in row.tolist() if str(clean).strip().lower() not in ("nan", "none", ""))
            if "cdp" in row_text or "custom design part" in row_text:
                current_section = "CDP"
                continue
            if "main pcb" in row_text or "display pcb" in row_text:
                current_section = "BOP"
                continue
            if current_section == "CDP":
                continue
            item_type = str(row[tc]).strip().upper() if tc else ''
            if item_type == 'CDP':
                continue
            mpn  = str(row[mc]).strip().split('\n')[0].split('/')[0].strip() if mc else ''
            desc = str(row[dc]).strip() if dc else ''
            qty  = row[qc] if qc else 1
            if _is_dnp_row(row_text, mpn, desc, qty, item_type):
                continue

            if not mpn or mpn.lower() in ('nan','none','nm','n/a','') or len(mpn) < 3: continue
            if not any(c.isdigit() for c in mpn) or not any(c.isalpha() for c in mpn): continue
            if str(qty).strip().upper() in ('NM','N/A',''): continue
            if mpn.upper() in seen: continue
            seen.add(mpn.upper())

            try: qty_val = int(float(str(qty)))
            except: qty_val = 1
            if qty_val <= 0: qty_val = 1

            rows.append({'MPN': mpn, 'Description': desc if desc and desc.lower() != 'nan' else '', 'Quantity': qty_val})

    except Exception as e:
        log.warning("BOM parse error: %s", e)

    # Build clean Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ BOM"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value     = f"REQUEST FOR QUOTATION — {rfq_code}"
    t.font      = Font(bold=True, size=13, color="1E3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill      = PatternFill("solid", fgColor="EBF3FF")
    ws.row_dimensions[1].height = 28

    # Instruction row
    ws.merge_cells('A2:H2')
    n = ws['A2']
    n.value     = "Please fill Unit Price, Stock, Lead Time and MOQ. Reply with completed sheet."
    n.font      = Font(italic=True, size=10, color="475569")
    n.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Headers
    headers    = ["#", "MPN / Part Number", "Description", "Quantity", "Unit Price (INR)", "Stock", "Lead Time", "MOQ"]
    col_widths = [5,   22,                  38,            12,         18,                  10,      14,          10]
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill      = hdr_fill
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 22

    # Data
    for ri, row in enumerate(rows, 4):
        fill = PatternFill("solid", fgColor=("F8FAFC" if ri % 2 == 0 else "FFFFFF"))
        data = [ri-3, row['MPN'], row['Description'], row['Quantity'], "", "", "", ""]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci==3))
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_message(rfq_code, vendor_name, vendor_email, bom_file_path):
    msg = MIMEMultipart("mixed")
    msg["From"]     = f"Procurement Team <{_SENDER}>"
    msg["To"]       = vendor_email
    msg["Subject"]  = f"{rfq_code} — Request for Quotation"
    msg["Reply-To"] = _SENDER

    plain = f"""Dear {vendor_name},

Please find the attached Bill of Materials for our upcoming project.

We request your quotation for each item:
  • Unit price (INR)
  • Available stock
  • Lead time
  • MOQ (if applicable)

Reference: {rfq_code}

Please reply with the completed sheet keeping the subject line intact.

Thank you,
Procurement Team | Strenth.ai"""

    html = f"""<html><body style="font-family:Arial,sans-serif;color:#1a1a1a;line-height:1.6;">
<p>Dear <strong>{vendor_name}</strong>,</p>
<p>Please find the attached Bill of Materials for our upcoming project.</p>
<p>We request your quotation for each item:</p>
<ul>
  <li>Unit price (INR)</li>
  <li>Available stock quantity</li>
  <li>Lead time</li>
  <li>MOQ (if applicable)</li>
</ul>
<p><strong>Reference:</strong> {rfq_code}</p>
<p>Please reply with the completed sheet keeping the subject line intact.</p>
<p>Thank you,<br><strong>Procurement Team</strong><br>Strenth.ai</p>
</body></html>"""

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html,  "html",  "utf-8"))
    msg.attach(alt)

    # Attach CLEAN BOM (not original file)
    try:
        bom_bytes = _generate_clean_bom(bom_file_path, rfq_code)
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(bom_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="RFQ_{rfq_code}.xlsx"')
        msg.attach(part)
    except Exception as e:
        log.error("Clean BOM generation failed: %s", e)

    return msg


def send_rfq_email(vendor_email: str, vendor_name: str, bom_file_path: str, rfq_code: str = "RFQ") -> bool:
    subject = f"{rfq_code} — Request for Quotation"
    plain = f"""Dear {vendor_name},\n\nPlease find the attached BOM for quotation.\nReference: {rfq_code}\n\nThank you,\nProcurement Team | Strenth.ai"""
    html = f"""<html><body style="font-family:Arial,sans-serif;"><p>Dear <strong>{vendor_name}</strong>,</p><p>Please find the attached BOM for quotation.</p><p><strong>Reference:</strong> {rfq_code}</p><p>Thank you,<br><strong>Procurement Team</strong><br>Strenth.ai</p></body></html>"""

    try:
        bom_bytes = _generate_clean_bom(bom_file_path, rfq_code)
    except Exception as e:
        log.error("Clean BOM generation failed: %s", e)
        bom_bytes = None

    if _ACS_CONN:
        return _send_via_acs(vendor_email, vendor_name, subject, plain, html,
                             bom_bytes, f"RFQ_{rfq_code}.xlsx")

    if not _SENDER or not _PASSWORD:
        log.error("No email provider configured. Set AZURE_COMM_CONNECTION_STRING or EMAIL_ADDRESS+EMAIL_PASSWORD")
        return False
    try:
        msg = _build_message(rfq_code, vendor_name, vendor_email, bom_file_path)
        with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT, timeout=15) as server:
            server.ehlo(); server.starttls(); server.ehlo()
            server.login(_SENDER, _PASSWORD)
            server.sendmail(_SENDER, vendor_email, msg.as_string())
        log.info("RFQ sent -> %s (%s)", vendor_name, vendor_email)
        return True
    except Exception as e:
        log.error("Send failed -> %s: %s", vendor_email, e)
        return False


def _normalize_cdp_rows(cdp_df) -> list[dict]:
    if cdp_df is None:
        return []
    if hasattr(cdp_df, "to_dict"):
        rows = cdp_df.to_dict(orient="records")
    else:
        rows = list(cdp_df)

    normalized = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        part_name = (
            row.get("part_name") or row.get("Part Name") or row.get("PartName") or
            row.get("name") or row.get("Name") or row.get("MPN") or row.get("mpn") or
            row.get("part") or row.get("Part") or ""
        )
        desc = row.get("description") or row.get("Description") or row.get("desc") or row.get("Desc") or ""
        qty = row.get("quantity") or row.get("Quantity") or row.get("qty") or row.get("Qty") or 1
        try:
            qty = int(float(str(qty).strip()))
        except Exception:
            qty = 1
        part_name = str(part_name).strip()
        if not part_name:
            continue
        normalized.append({
            "part_name": part_name,
            "description": str(desc).strip(),
            "quantity": max(qty, 1),
            "note": "Custom manufacturing required",
        })
    return normalized


def _generate_cdp_bom(cdp_rows: list[dict], rfq_code: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CDP RFQ"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:E1')
    title = ws['A1']
    title.value = f"CDP MANUFACTURING RFQ - {rfq_code}"
    title.font = Font(bold=True, size=13, color="1E3A5F")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor="EBF3FF")

    headers = ["#", "Part Name", "Description", "Quantity", "Note"]
    widths = [5, 28, 44, 12, 32]
    fill = PatternFill("solid", fgColor="1E3A5F")
    for ci, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for ri, row in enumerate(cdp_rows, 3):
        data = [ri - 2, row["part_name"], row["description"], row["quantity"], row["note"]]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in (3, 5)))

    ws.freeze_panes = 'A3'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_cdp_message(rfq_code: str, cm_name: str, cm_email: str, cdp_rows: list[dict]):
    msg = MIMEMultipart("mixed")
    msg["From"] = f"Procurement Team <{_SENDER}>"
    msg["To"] = cm_email
    msg["Subject"] = f"{rfq_code} - CDP Manufacturing RFQ"
    msg["Reply-To"] = _SENDER

    plain_lines = "\n".join(
        f"- {r['part_name']} | {r['description']} | Qty: {r['quantity']} | Custom manufacturing required"
        for r in cdp_rows
    )
    plain = f"""Dear {cm_name},

Please quote for the attached CDP items.

{plain_lines}

Note: Custom manufacturing required
Reference: {rfq_code}

Thank you,
Procurement Team | Strenth.ai"""

    html_rows = "".join(
        f"<tr><td>{r['part_name']}</td><td>{r['description']}</td><td>{r['quantity']}</td><td>{r['note']}</td></tr>"
        for r in cdp_rows
    )
    html = f"""<html><body style="font-family:Arial,sans-serif;color:#1a1a1a;line-height:1.6;">
<p>Dear <strong>{cm_name}</strong>,</p>
<p>Please quote for the attached CDP items.</p>
<table border="1" cellspacing="0" cellpadding="6">
<thead><tr><th>Part Name</th><th>Description</th><th>Quantity</th><th>Note</th></tr></thead>
<tbody>{html_rows}</tbody>
</table>
<p><strong>Note:</strong> Custom manufacturing required</p>
<p><strong>Reference:</strong> {rfq_code}</p>
<p>Thank you,<br><strong>Procurement Team</strong><br>Strenth.ai</p>
</body></html>"""

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(alt)

    cdp_bytes = _generate_cdp_bom(cdp_rows, rfq_code)
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(cdp_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="CDP_RFQ_{rfq_code}.xlsx"')
    msg.attach(part)
    return msg


def send_cdp_rfq(cdp_df, contract_manufacturers: list[dict] | None = None, rfq_code: str = "CDP-RFQ") -> int:
    cdp_rows = _normalize_cdp_rows(cdp_df)
    if not cdp_rows:
        return 0
    if not _SENDER or not _PASSWORD:
        log.error("EMAIL_ADDRESS or EMAIL_PASSWORD not set")
        return 0

    recipients = contract_manufacturers if contract_manufacturers is not None else get_all_contract_manufacturers()
    sent = 0
    for cm in recipients:
        email = cm.get("cm_email") or cm.get("email") or ""
        name = cm.get("cm_name") or cm.get("name") or "Contract Manufacturer"
        if not email:
            continue
        try:
            msg = _build_cdp_message(rfq_code, name, email, cdp_rows)
            with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT, timeout=15) as server:
                server.ehlo(); server.starttls(); server.ehlo()
                server.login(_SENDER, _PASSWORD)
                server.sendmail(_SENDER, email, msg.as_string())
            sent += 1
            log.info("CDP RFQ sent -> %s (%s)", name, email)
        except Exception as e:
            log.error("CDP RFQ send failed -> %s: %s", email, e)
    return sent


def _build_mechanical_message(rfq_code: str, cm_name: str, cm_email: str, rows: list[dict]):
    msg = MIMEMultipart("mixed")
    msg["From"] = f"Procurement Team <{_SENDER}>"
    msg["To"] = cm_email
    msg["Subject"] = f"{rfq_code} - Mechanical Manufacturing RFQ"
    msg["Reply-To"] = _SENDER

    for row in rows:
        row["note"] = "Mechanical manufacturing required"

    plain_lines = "\n".join(
        f"- {r['part_name']} | {r['description']} | Qty: {r['quantity']} | Mechanical manufacturing required"
        for r in rows
    )
    plain = f"""Dear {cm_name},

Please quote for the attached mechanical parts.

{plain_lines}

Note: Mechanical manufacturing required
Reference: {rfq_code}

Thank you,
Procurement Team | Strenth.ai"""

    html_rows = "".join(
        f"<tr><td>{r['part_name']}</td><td>{r['description']}</td><td>{r['quantity']}</td><td>{r['note']}</td></tr>"
        for r in rows
    )
    html = f"""<html><body style="font-family:Arial,sans-serif;color:#1a1a1a;line-height:1.6;">
<p>Dear <strong>{cm_name}</strong>,</p>
<p>Please quote for the attached mechanical parts.</p>
<table border="1" cellspacing="0" cellpadding="6">
<thead><tr><th>Part Name</th><th>Description</th><th>Quantity</th><th>Note</th></tr></thead>
<tbody>{html_rows}</tbody>
</table>
<p><strong>Note:</strong> Mechanical manufacturing required</p>
<p><strong>Reference:</strong> {rfq_code}</p>
<p>Thank you,<br><strong>Procurement Team</strong><br>Strenth.ai</p>
</body></html>"""

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(alt)

    mech_bytes = _generate_cdp_bom(rows, rfq_code)
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(mech_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="MECHANICAL_RFQ_{rfq_code}.xlsx"')
    msg.attach(part)
    return msg


def send_mechanical_rfq(mechanical_df, contract_manufacturers: list[dict] | None = None, rfq_code: str = "MECH-RFQ") -> int:
    rows = _normalize_cdp_rows(mechanical_df)
    if not rows:
        return 0
    if not _SENDER or not _PASSWORD:
        log.error("EMAIL_ADDRESS or EMAIL_PASSWORD not set")
        return 0

    recipients = contract_manufacturers if contract_manufacturers is not None else get_all_contract_manufacturers()
    sent = 0
    for cm in recipients:
        email = cm.get("cm_email") or cm.get("email") or ""
        name = cm.get("cm_name") or cm.get("name") or "Contract Manufacturer"
        if not email:
            continue
        try:
            msg = _build_mechanical_message(rfq_code, name, email, rows)
            with smtplib.SMTP(_SMTP_HOST, _SMTP_PORT, timeout=15) as server:
                server.ehlo(); server.starttls(); server.ehlo()
                server.login(_SENDER, _PASSWORD)
                server.sendmail(_SENDER, email, msg.as_string())
            sent += 1
            log.info("Mechanical RFQ sent -> %s (%s)", name, email)
        except Exception as e:
            log.error("Mechanical RFQ send failed -> %s: %s", email, e)
    return sent


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    send_rfq_email(_SENDER, "Test Vendor", "test.xlsx", "RFQ-TEST1")
